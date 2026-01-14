#!/usr/bin/env python3
"""
FastAPI Service Wrapper for Excel Sync Agent
===========================================
This service exposes the Excel Sync Agent functionality via HTTP API endpoints.
It can run as a standalone service or be integrated with the main backend.

Usage:
    python agent_api.py --config config.json --port 8001
"""

import json
import logging
import os
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

from fastapi import FastAPI, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

# Import the agent classes from local_sync_agent
from local_sync_agent import SyncAgent, ExcelReader, load_config
import openpyxl
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('agent_api.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('AgentAPI')

# Global agent instance
_agent_instance: Optional[SyncAgent] = None
_agent_config: Optional[Dict] = None
_sync_history: List[Dict[str, Any]] = []
_sync_status: Dict[str, Any] = {
    "status": "stopped",
    "last_sync": None,
    "active_projects": [],
    "errors": []
}

# FastAPI app
app = FastAPI(
    title="Excel Sync Agent API",
    description="HTTP API for Excel Sync Agent",
    version="1.0.0"
)

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================
# REQUEST/RESPONSE MODELS
# ============================================================================

class SyncTriggerRequest(BaseModel):
    project_id: Optional[str] = None  # If None, sync all projects
    force: bool = False

class SyncStatusResponse(BaseModel):
    status: str
    last_sync: Optional[str]
    active_projects: List[str]
    errors: List[str]
    agent_configured: bool

class ProjectInfo(BaseModel):
    project_id: str
    project_name: str
    excel_file: str
    sheet_name: str
    cell_mappings: Dict[str, str]
    last_synced: Optional[str] = None
    file_exists: bool = False

class SyncResult(BaseModel):
    success: bool
    project_id: str
    message: str
    data: Optional[Dict[str, Any]] = None
    timestamp: str

# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def get_agent() -> SyncAgent:
    """Get or initialize the agent instance"""
    global _agent_instance, _agent_config
    
    if _agent_instance is None:
        if _agent_config is None:
            raise HTTPException(
                status_code=503,
                detail="Agent not configured. Call /api/agent/configure first."
            )
        _agent_instance = SyncAgent(_agent_config)
        logger.info("Agent instance initialized")
    
    return _agent_instance

def update_sync_status(status: str, project_id: Optional[str] = None, error: Optional[str] = None):
    """Update global sync status"""
    global _sync_status
    
    _sync_status["status"] = status
    _sync_status["last_sync"] = datetime.utcnow().isoformat()
    
    if project_id and project_id not in _sync_status["active_projects"]:
        _sync_status["active_projects"].append(project_id)
    
    if error:
        _sync_status["errors"].append({
            "timestamp": datetime.utcnow().isoformat(),
            "error": error,
            "project_id": project_id
        })
        # Keep only last 10 errors
        _sync_status["errors"] = _sync_status["errors"][-10:]

def add_sync_history(project_id: str, success: bool, message: str, data: Optional[Dict] = None):
    """Add entry to sync history"""
    global _sync_history
    
    entry = {
        "project_id": project_id,
        "success": success,
        "message": message,
        "data": data,
        "timestamp": datetime.utcnow().isoformat()
    }
    
    _sync_history.append(entry)
    # Keep only last 100 entries
    _sync_history = _sync_history[-100:]


def get_allowed_directories() -> List[Path]:
    """Get list of allowed directories from config"""
    if _agent_config is None:
        # Default to common safe directories
        return [
            Path.home() / "Desktop",
            Path.home() / "Documents",
        ]
    
    allowed = _agent_config.get("allowed_directories", [])
    if not allowed:
        # Default to common safe directories
        return [
            Path.home() / "Desktop",
            Path.home() / "Documents",
        ]
    
    # Expand user paths and convert to Path objects
    directories = []
    for dir_path in allowed:
        expanded = Path(dir_path).expanduser().resolve()
        if expanded.exists() and expanded.is_dir():
            directories.append(expanded)
    
    return directories if directories else [Path.home() / "Desktop"]


def validate_path(file_path: str) -> Path:
    """
    Validate that a file path is within allowed directories.
    Prevents directory traversal attacks.
    
    Returns:
        Resolved Path object if valid
        
    Raises:
        HTTPException if path is not allowed
    """
    try:
        requested_path = Path(file_path).expanduser().resolve()
        allowed_dirs = get_allowed_directories()
        
        # Check if path is within any allowed directory
        is_allowed = False
        for allowed_dir in allowed_dirs:
            try:
                # Check if requested_path is within allowed_dir
                requested_path.relative_to(allowed_dir)
                is_allowed = True
                break
            except ValueError:
                # Path is not relative to this allowed directory
                continue
        
        if not is_allowed:
            raise HTTPException(
                status_code=403,
                detail=f"Access denied: Path '{file_path}' is not within allowed directories. Allowed: {[str(d) for d in allowed_dirs]}"
            )
        
        return requested_path
    
    except Exception as e:
        if isinstance(e, HTTPException):
            raise
        raise HTTPException(
            status_code=400,
            detail=f"Invalid path: {str(e)}"
        )

# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/")
async def root():
    """Root endpoint"""
    return {
        "service": "Excel Sync Agent API",
        "version": "1.0.0",
        "status": "running",
        "agent_configured": _agent_instance is not None,
        "endpoints": {
            "health": "/health",
            "status": "/api/agent/status",
            "projects": "/api/agent/projects",
            "sync": "/api/agent/sync",
            "history": "/api/agent/history",
            "configure": "/api/agent/configure",
            "files_list": "/api/agent/files/list",
            "excel_info": "/api/agent/files/excel/info",
            "excel_read": "/api/agent/files/excel/read"
        }
    }

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.utcnow().isoformat(),
        "agent_configured": _agent_instance is not None,
        "config_loaded": _agent_config is not None
    }

@app.post("/api/agent/configure")
async def configure_agent(config: Dict[str, Any]):
    """
    Configure the agent with a new configuration.
    This allows dynamic configuration without restarting the service.
    """
    global _agent_instance, _agent_config
    
    try:
        # Validate config structure
        required_keys = ["platform_url", "api_key", "projects"]
        for key in required_keys:
            if key not in config:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing required config key: {key}"
                )
        
        # Validate projects
        if not isinstance(config["projects"], list) or len(config["projects"]) == 0:
            raise HTTPException(
                status_code=400,
                detail="Config must contain at least one project"
            )
        
        # Store config and reset agent instance
        _agent_config = config
        _agent_instance = None  # Will be recreated on next use
        
        logger.info(f"Agent configured with {len(config['projects'])} projects")
        
        return {
            "status": "configured",
            "projects_count": len(config["projects"]),
            "message": "Agent configuration updated successfully"
        }
    
    except Exception as e:
        logger.error(f"Configuration error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/status", response_model=SyncStatusResponse)
async def get_sync_status():
    """Get current sync status"""
    return SyncStatusResponse(
        status=_sync_status["status"],
        last_sync=_sync_status["last_sync"],
        active_projects=_sync_status["active_projects"],
        errors=[e.get("error", "") for e in _sync_status["errors"]],
        agent_configured=_agent_instance is not None
    )

@app.get("/api/agent/projects", response_model=List[ProjectInfo])
async def get_projects():
    """Get list of all configured projects with their status"""
    if _agent_config is None:
        raise HTTPException(
            status_code=503,
            detail="Agent not configured"
        )
    
    projects = []
    for project_config in _agent_config.get("projects", []):
        excel_path = Path(project_config.get("excel_file", ""))
        
        # Find last sync for this project
        last_synced = None
        for entry in reversed(_sync_history):
            if entry.get("project_id") == project_config.get("project_id"):
                if entry.get("success"):
                    last_synced = entry.get("timestamp")
                    break
        
        projects.append(ProjectInfo(
            project_id=project_config.get("project_id", ""),
            project_name=project_config.get("project_name", ""),
            excel_file=str(excel_path),
            sheet_name=project_config.get("sheet_name", ""),
            cell_mappings=project_config.get("cell_mappings", {}),
            last_synced=last_synced,
            file_exists=excel_path.exists()
        ))
    
    return projects

@app.post("/api/agent/sync", response_model=SyncResult)
async def trigger_sync(request: SyncTriggerRequest, background_tasks: BackgroundTasks):
    """
    Trigger a sync for a specific project or all projects.
    Returns immediately and syncs in the background.
    """
    if _agent_config is None:
        raise HTTPException(
            status_code=503,
            detail="Agent not configured"
        )
    
    agent = get_agent()
    
    if request.project_id:
        # Sync specific project
        project_config = None
        for proj in _agent_config.get("projects", []):
            if proj.get("project_id") == request.project_id:
                project_config = proj
                break
        
        if not project_config:
            raise HTTPException(
                status_code=404,
                detail=f"Project '{request.project_id}' not found"
            )
        
        # Run sync in background
        background_tasks.add_task(sync_project_task, project_config)
        
        return SyncResult(
            success=True,
            project_id=request.project_id,
            message=f"Sync initiated for project {request.project_id}",
            timestamp=datetime.utcnow().isoformat()
        )
    else:
        # Sync all projects
        background_tasks.add_task(sync_all_projects_task)
        
        return SyncResult(
            success=True,
            project_id="all",
            message="Sync initiated for all projects",
            timestamp=datetime.utcnow().isoformat()
        )

def sync_project_task(project_config: Dict):
    """Background task to sync a single project"""
    project_id = project_config.get("project_id", "unknown")
    
    try:
        update_sync_status("syncing", project_id)
        logger.info(f"Starting sync for project {project_id}")
        
        agent = get_agent()
        success = agent.sync_project(project_config)
        
        if success:
            # Read the data that was synced
            reader = ExcelReader(project_config.get("excel_file"))
            data = reader.read_cells(
                project_config.get("sheet_name"),
                project_config.get("cell_mappings", {})
            )
            
            add_sync_history(
                project_id,
                True,
                f"Successfully synced project {project_id}",
                data
            )
            update_sync_status("idle", project_id)
            logger.info(f"Successfully synced project {project_id}")
        else:
            add_sync_history(
                project_id,
                False,
                f"Failed to sync project {project_id}"
            )
            update_sync_status("idle", project_id, f"Sync failed for {project_id}")
            logger.error(f"Failed to sync project {project_id}")
    
    except Exception as e:
        error_msg = f"Error syncing project {project_id}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        add_sync_history(project_id, False, error_msg)
        update_sync_status("idle", project_id, error_msg)

def sync_all_projects_task():
    """Background task to sync all projects"""
    try:
        update_sync_status("syncing")
        logger.info("Starting sync for all projects")
        
        agent = get_agent()
        agent.sync_all_projects()
        
        update_sync_status("idle")
        logger.info("Completed sync for all projects")
    
    except Exception as e:
        error_msg = f"Error syncing all projects: {str(e)}"
        logger.error(error_msg, exc_info=True)
        update_sync_status("idle", error=error_msg)

@app.get("/api/agent/history")
async def get_sync_history(limit: int = 50):
    """Get sync history"""
    return {
        "history": _sync_history[-limit:],
        "total": len(_sync_history)
    }

@app.get("/api/agent/project/{project_id}/data")
async def get_project_data(project_id: str):
    """Get current data from a project's Excel file without syncing"""
    if _agent_config is None:
        raise HTTPException(
            status_code=503,
            detail="Agent not configured"
        )
    
    project_config = None
    for proj in _agent_config.get("projects", []):
        if proj.get("project_id") == project_id:
            project_config = proj
            break
    
    if not project_config:
        raise HTTPException(
            status_code=404,
            detail=f"Project '{project_id}' not found"
        )
    
    excel_path = Path(project_config.get("excel_file", ""))
    if not excel_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Excel file not found: {excel_path}"
        )
    
    try:
        reader = ExcelReader(str(excel_path))
        data = reader.read_cells(
            project_config.get("sheet_name"),
            project_config.get("cell_mappings", {})
        )
        
        return {
            "project_id": project_id,
            "data": data,
            "timestamp": datetime.utcnow().isoformat()
        }
    
    except Exception as e:
        logger.error(f"Error reading project data: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error reading Excel file: {str(e)}"
        )

# ============================================================================
# FILE OPERATION ENDPOINTS (For Cloud Agent Tools)
# ============================================================================

@app.get("/api/agent/files/list")
async def list_files(directory: str = None):
    """
    List files in a directory on the user's local machine.
    Directory must be within allowed directories configured in agent config.
    
    Args:
        directory: Directory path to list (default: user's Desktop)
    
    Returns:
        List of files and directories
    """
    try:
        if directory is None:
            directory = str(Path.home() / "Desktop")
        
        # Validate path is within allowed directories
        dir_path = validate_path(directory)
        
        if not dir_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Directory not found: {directory}"
            )
        
        if not dir_path.is_dir():
            raise HTTPException(
                status_code=400,
                detail=f"Path is not a directory: {directory}"
            )
        
        # List files and directories
        items = []
        for item in dir_path.iterdir():
            try:
                stat = item.stat()
                items.append({
                    "name": item.name,
                    "path": str(item),
                    "is_directory": item.is_dir(),
                    "size_bytes": stat.st_size if item.is_file() else None,
                    "modified_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
                })
            except Exception as e:
                logger.warning(f"Could not stat {item}: {e}")
                continue
        
        # Sort: directories first, then by name
        items.sort(key=lambda x: (not x["is_directory"], x["name"].lower()))
        
        return {
            "directory": str(dir_path),
            "count": len(items),
            "items": items
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error listing files: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error listing directory: {str(e)}"
        )


@app.get("/api/agent/files/excel/info")
async def get_excel_info(file_path: str):
    """
    Get structure and metadata about an Excel file.
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Excel file structure: sheets, columns, dimensions
    """
    try:
        # Validate path
        excel_path = validate_path(file_path)
        
        if not excel_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found: {file_path}"
            )
        
        if not excel_path.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
            raise HTTPException(
                status_code=400,
                detail=f"File is not an Excel file: {file_path}"
            )
        
        # Read Excel structure
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=False, read_only=True)
            
            sheets_info = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get dimensions
                max_row = ws.max_row
                max_col = ws.max_column
                
                # Sample first row for column names
                columns = []
                if max_row > 0:
                    for col in range(1, min(max_col + 1, 11)):  # Limit to first 10 columns
                        cell = ws.cell(1, col)
                        if cell.value:
                            columns.append({
                                "column": get_column_letter(col),
                                "header": str(cell.value)
                            })
                
                # Count formulas - safe access for EmptyCell
                formula_count = 0
                for row in ws.iter_rows(min_row=1, max_row=min(max_row, 100), max_col=min(max_col, 20)):
                    for cell in row:
                        cell_data_type = getattr(cell, 'data_type', None)
                        if cell_data_type == 'f':  # Formula
                            formula_count += 1
                
                sheets_info.append({
                    "sheet_name": sheet_name,
                    "dimensions": {
                        "max_row": max_row,
                        "max_column": max_col
                    },
                    "columns": columns,
                    "formula_count": formula_count,
                    "has_data": max_row > 0
                })
            
            wb.close()
            
            return {
                "file_path": str(excel_path),
                "file_name": excel_path.name,
                "total_sheets": len(sheets_info),
                "sheets": sheets_info,
                "timestamp": datetime.utcnow().isoformat()
            }
        
        except Exception as e:
            logger.error(f"Error reading Excel file structure: {e}")
            raise HTTPException(
                status_code=500,
                detail=f"Error reading Excel file: {str(e)}"
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting Excel info: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error: {str(e)}"
        )


@app.get("/api/agent/files/excel/read")
async def read_excel_file(file_path: str, sheet_name: str = None, max_rows: int = 100, include_formulas: bool = True, include_code_refs: bool = True):
    """
    Read data from an Excel file with enhanced context.
    
    This endpoint provides:
    - Cell values (calculated results)
    - Formulas (if include_formulas=True)
    - Building code references (if include_code_refs=True)
    - Structured context: label-value pairs for better understanding
    
    Args:
        file_path: Path to Excel file
        sheet_name: Optional sheet name (reads first sheet if not specified)
        max_rows: Maximum number of rows to read (default: 100)
        include_formulas: Include formulas in response (default: True)
        include_code_refs: Extract building code references (default: True)
    
    Returns:
        Excel data with formulas, code references, and structured context
    """
    try:
        # Validate path
        excel_path = validate_path(file_path)
        
        if not excel_path.exists():
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found: {file_path}"
            )
        
        if not excel_path.suffix.lower() in ['.xlsx', '.xlsm', '.xls']:
            raise HTTPException(
                status_code=400,
                detail=f"File is not an Excel file: {file_path}"
            )
        
        # Read Excel data - use data_only=False to get formulas if needed
        read_formulas = include_formulas or include_code_refs
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=not read_formulas, read_only=True)
            
            # Use first sheet if not specified
            if sheet_name is None:
                sheet_name = wb.sheetnames[0]
            
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise HTTPException(
                    status_code=404,
                    detail=f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
                )
            
            ws = wb[sheet_name]
            
            # Read data with enhanced context
            data_rows = []
            headers = []
            formulas_map = {}  # cell_address -> formula with dependencies
            formulas_detailed = []  # List of formula objects with full context
            code_references = []
            structured_context = []  # Optional label-value pairs (flexible, not required)
            
            import re
            code_reference_pattern = re.compile(
                r'(?:CSA|NBC|OBC|ASCE|AISC|AISI|NDS|CSA\s+O86|CSA\s+S16|CSA\s+A23\.3)[\s-]*(?:O86|S16|A23\.3|NBC|OBC)?[\s-]*(?:-?\d{2,4})?[\s,]*[Cc]l(?:ause)?[\s.]*(\d+(?:\.\d+)*(?:\.\d+)?)',
                re.IGNORECASE
            )
            
            def parse_formula_dependencies(formula: str) -> dict:
                """
                Parse Excel formula to extract cell references and dependencies.
                Handles: A1, $A$1, A$1, $A1, Sheet1!A1, A1:B5, Sheet1!A1:B5, etc.
                Returns dict with:
                - formula: Original formula string
                - dependencies: List of cell references (e.g., ['B5', 'B6', 'C10'])
                - dependency_count: Number of dependencies
                """
                if not formula or not isinstance(formula, str) or not formula.startswith('='):
                    return {"formula": formula, "dependencies": [], "dependency_count": 0}
                
                dependencies = set()  # Use set to avoid duplicates
                
                # Pattern 1: Simple cell references (A1, $A$1, A$1, $A1)
                # Matches: column letters (1-3 digits) + row number
                simple_cell_pattern = re.compile(r'\$?[A-Z]{1,3}\$?\d+')
                
                # Pattern 2: Sheet-qualified references (Sheet1!A1, 'Sheet Name'!A1)
                sheet_cell_pattern = re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_]*)!\$?[A-Z]{1,3}\$?\d+")
                
                # Pattern 3: Ranges (A1:B5, Sheet1!A1:B5)
                range_pattern = re.compile(r"(?:'[^']+'|[A-Za-z_][A-Za-z0-9_]*)!\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+|\$?[A-Z]{1,3}\$?\d+:\$?[A-Z]{1,3}\$?\d+")
                
                # Extract ranges first (to handle them specially)
                range_matches = range_pattern.findall(formula)
                for match in range_matches:
                    # Extract start and end cells from range
                    if '!' in match:
                        # Sheet-qualified range: Sheet1!A1:B5
                        cell_part = match.split('!')[-1]
                    else:
                        # Simple range: A1:B5
                        cell_part = match
                    
                    if ':' in cell_part:
                        start_cell, end_cell = cell_part.split(':')
                        # Remove $ signs
                        start_cell = start_cell.replace('$', '')
                        end_cell = end_cell.replace('$', '')
                        dependencies.add(start_cell)
                        dependencies.add(end_cell)
                
                # Extract sheet-qualified cells (not in ranges)
                sheet_matches = sheet_cell_pattern.findall(formula)
                for match in sheet_matches:
                    if ':' not in match:  # Skip if it's part of a range (already handled)
                        cell = match.split('!')[-1].replace('$', '')
                        if cell and ':' not in cell:  # Make sure it's not a range
                            dependencies.add(cell)
                
                # Extract simple cells (not in ranges, not sheet-qualified)
                # Remove ranges and sheet-qualified refs from formula first
                temp_formula = formula
                for range_match in range_pattern.findall(formula):
                    temp_formula = temp_formula.replace(range_match, '')
                for sheet_match in sheet_cell_pattern.findall(formula):
                    temp_formula = temp_formula.replace(sheet_match, '')
                
                simple_matches = simple_cell_pattern.findall(temp_formula)
                for match in simple_matches:
                    if ':' not in match:  # Skip if it's part of a range
                        cell = match.replace('$', '')
                        if cell:
                            dependencies.add(cell)
                
                return {
                    "formula": formula,
                    "dependencies": sorted(list(dependencies)),
                    "dependency_count": len(dependencies)
                }
            
            # Get headers from first row
            if ws.max_row > 0:
                for col in range(1, min(ws.max_column + 1, 51)):  # Limit to 50 columns
                    cell = ws.cell(1, col)
                    headers.append(cell.value if cell.value is not None else f"Column{col}")
            
            # Read data rows with formulas and context
            for row_idx in range(2, min(ws.max_row + 1, max_rows + 2)):  # Start from row 2, limit rows
                row_data = {}
                row_formulas = {}
                row_labels = {}  # Track labels in this row for context
                
                for col_idx, header in enumerate(headers, start=1):
                    if col_idx > ws.max_column:
                        break
                    cell = ws.cell(row_idx, col_idx)
                    # Construct cell address manually to handle EmptyCell objects
                    cell_address = f"{get_column_letter(col_idx)}{row_idx}"
                    
                    # Get cell value - safe access for EmptyCell
                    cell_value = getattr(cell, 'value', None)
                    row_data[header] = cell_value
                    
                    # Extract formulas if requested - safe access for EmptyCell
                    cell_data_type = getattr(cell, 'data_type', None)
                    if include_formulas and cell_data_type == 'f' and cell_value:
                        formula_str = str(cell_value)
                        # Parse formula to get dependencies
                        formula_info = parse_formula_dependencies(formula_str)
                        formula_info.update({
                            "cell": cell_address,
                            "row": row_idx,
                            "column": get_column_letter(col_idx),
                            "calculated_value": cell_value
                        })
                        
                        formulas_map[cell_address] = formula_str  # Keep simple map for backward compat
                        formulas_detailed.append(formula_info)  # Detailed info with dependencies
                        row_formulas[header] = formula_str
                    
                    # Extract code references if requested
                    if include_code_refs:
                        # Check cell value for code references
                        if cell_value:
                            cell_text = str(cell_value)
                            matches = code_reference_pattern.findall(cell_text)
                            for match in matches:
                                code_references.append({
                                    "cell": cell_address,
                                    "reference": cell_text[:200],
                                    "clause": match,
                                    "sheet": sheet_name,
                                    "row": row_idx
                                })
                        
                        # Check cell comments for code references - safe access for EmptyCell
                        cell_comment = getattr(cell, 'comment', None)
                        if cell_comment:
                            comment_text = getattr(cell_comment, 'text', '')
                            if comment_text:
                                matches = code_reference_pattern.findall(comment_text)
                                for match in matches:
                                    code_references.append({
                                        "cell": cell_address,
                                        "reference": comment_text[:200],
                                        "clause": match,
                                        "sheet": sheet_name,
                                        "row": row_idx,
                                        "source": "comment"
                                    })
                    
                    # OPTIONAL: Build structured context - flexible pattern detection
                    # This is a best-effort attempt to identify label-value pairs
                    # If patterns don't match, that's OK - the LLM will figure it out from raw data
                    # We try multiple patterns but don't require any to succeed
                    
                    # Pattern 1: Label in left column, value in right column (common form layout)
                    if cell_value is not None and isinstance(cell_value, (int, float)):
                        if col_idx > 1:
                            try:
                                left_cell = ws.cell(row_idx, col_idx - 1)
                                left_cell_value = getattr(left_cell, 'value', None)
                                if left_cell_value and isinstance(left_cell_value, str):
                                    label = str(left_cell_value).strip()
                                    # Heuristic: likely a label if it's short, not a formula, and not a number
                                    if (label and 2 <= len(label) <= 80 and 
                                        not label.startswith("=") and 
                                        not label.replace('.', '').replace('-', '').isdigit()):
                                        # Construct left cell address manually
                                        left_cell_address = f"{get_column_letter(col_idx - 1)}{row_idx}"
                                        row_labels[label] = {
                                            "value": cell_value,
                                            "cell": cell_address,
                                            "label_cell": left_cell_address,
                                            "type": "left_label"
                                        }
                            except:
                                pass  # Ignore errors - pattern detection is optional
                    
                    # Pattern 2: Inline label:value pattern
                    if isinstance(cell_value, str) and ":" in cell_value:
                        parts = cell_value.split(":", 1)
                        if len(parts) == 2:
                            potential_label = parts[0].strip()
                            potential_value = parts[1].strip()
                            if potential_label and potential_value and len(potential_label) < 80:
                                row_labels[potential_label] = {
                                    "value": potential_value,
                                    "cell": cell_address,
                                    "type": "inline"
                                }
                
                # Add formulas to row data if present
                if row_formulas:
                    row_data["_formulas"] = row_formulas
                
                # Add structured context (label-value pairs) if found
                if row_labels:
                    structured_context.append({
                        "row": row_idx,
                        "type": "vertical",
                        "labels": row_labels
                    })
                
                # Pattern 3: HORIZONTAL PATTERN - Related cells in same row across columns
                # Detects patterns like: "Load Duration Factor," "KD", "=", 1, "comment..."
                # This identifies when a label, variable name, value, and comment are in the same row
                # Only check once per row (not already processed)
                if not any(ctx.get("row") == row_idx and ctx.get("type") == "horizontal" for ctx in structured_context):
                    try:
                        # Scan the entire row for horizontal patterns
                        row_cells = []
                        for scan_col in range(1, min(ws.max_column + 1, 51)):
                            scan_cell = ws.cell(row_idx, scan_col)
                            scan_value = getattr(scan_cell, 'value', None)
                            if scan_value is not None:
                                row_cells.append({
                                    "col": scan_col,
                                    "value": scan_value,
                                    "cell": f"{get_column_letter(scan_col)}{row_idx}"
                                })
                        
                        # Detect horizontal pattern: label, variable, =, value, comment
                        # Pattern: text label, short text (variable), "=" or number, number, long text (comment)
                        if len(row_cells) >= 3:
                            horizontal_relationship = None
                            
                            # Look for label (text, often ends with comma) followed by variable, value, comment
                            for i, cell_info in enumerate(row_cells):
                                value = cell_info["value"]
                                if isinstance(value, str) and len(value) > 5 and len(value) < 60:
                                    # Could be a label
                                    label_candidate = value.strip().rstrip(',')
                                    
                                    # Look ahead for variable name (short text, 1-5 chars)
                                    if i + 1 < len(row_cells):
                                        var_candidate = row_cells[i + 1]["value"]
                                        if isinstance(var_candidate, str) and 1 <= len(str(var_candidate)) <= 5:
                                            # Look ahead for "=" or number
                                            if i + 2 < len(row_cells):
                                                eq_or_val = row_cells[i + 2]["value"]
                                                # Look ahead for actual value (number)
                                                if i + 3 < len(row_cells):
                                                    val_candidate = row_cells[i + 3]["value"]
                                                    if isinstance(val_candidate, (int, float)):
                                                        # Found pattern: label, variable, =/symbol, value
                                                        horizontal_relationship = {
                                                            "label": label_candidate,
                                                            "variable": str(var_candidate),
                                                            "value": val_candidate,
                                                            "label_cell": cell_info["cell"],
                                                            "variable_cell": row_cells[i + 1]["cell"],
                                                            "value_cell": row_cells[i + 3]["cell"],
                                                            "row": row_idx
                                                        }
                                                        
                                                        # Look for comment (long text) in remaining cells
                                                        for j in range(i + 4, len(row_cells)):
                                                            comment_candidate = row_cells[j]["value"]
                                                            if isinstance(comment_candidate, str) and len(comment_candidate) > 20:
                                                                horizontal_relationship["comment"] = comment_candidate[:200]
                                                                horizontal_relationship["comment_cell"] = row_cells[j]["cell"]
                                                                break
                                                        
                                                        break
                            
                            if horizontal_relationship:
                                structured_context.append({
                                    "row": row_idx,
                                    "type": "horizontal",
                                    "relationship": horizontal_relationship
                                })
                    except Exception as e:
                        logger.debug(f"Horizontal pattern detection failed for row {row_idx}: {e}")
                        pass  # Ignore errors - pattern detection is optional
                
                # Only add row if it has at least one non-None value
                if any(v is not None for v in row_data.values() if v != "_formulas"):
                    data_rows.append(row_data)
            
            wb.close()
            
            result = {
                "file_path": str(excel_path),
                "sheet_name": sheet_name,
                "headers": headers,
                "row_count": len(data_rows),
                "data": data_rows,
                "timestamp": datetime.utcnow().isoformat()
            }
            
            # Add formulas if requested - provide both simple map and detailed info
            if include_formulas:
                if formulas_map:
                    result["formulas"] = formulas_map  # Simple map: cell -> formula string
                    result["formula_count"] = len(formulas_map)
                
                if formulas_detailed:
                    result["formulas_detailed"] = formulas_detailed  # Detailed: includes dependencies
                    # Build dependency graph summary
                    all_dependencies = set()
                    for f in formulas_detailed:
                        all_dependencies.update(f.get("dependencies", []))
                    result["formula_dependencies_summary"] = {
                        "total_unique_dependencies": len(all_dependencies),
                        "dependent_cells": sorted(list(all_dependencies))
                    }
            
            # Add code references if requested
            if include_code_refs and code_references:
                result["code_references"] = code_references
                result["code_reference_count"] = len(code_references)
            
            # Add structured context (label-value pairs) - OPTIONAL, may be empty
            # This is a best-effort attempt - if empty, LLM will infer from raw data
            if structured_context:
                result["structured_context"] = structured_context
                result["context_pairs_count"] = sum(len(ctx["labels"]) for ctx in structured_context)
            else:
                result["structured_context"] = []  # Explicitly empty - not an error
                result["context_pairs_count"] = 0
            
            return result
        
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise HTTPException(
                status_code=500,
                detail=f"Error reading Excel file: {str(e)}"
            )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error reading Excel: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"Error: {str(e)}"
        )

# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='Excel Sync Agent API Service')
    parser.add_argument('--config', help='Path to configuration JSON file')
    parser.add_argument('--port', type=int, default=8001, help='Port to run the API on')
    parser.add_argument('--host', default='0.0.0.0', help='Host to bind to')
    
    args = parser.parse_args()
    
    # Load config if provided
    if args.config:
        try:
            global _agent_config
            _agent_config = load_config(args.config)
            logger.info(f"Loaded configuration from {args.config}")
            logger.info(f"Configured with {len(_agent_config.get('projects', []))} projects")
        except Exception as e:
            logger.error(f"Failed to load config: {e}")
            logger.warning("Starting without configuration. Use /api/agent/configure to configure.")
    
    print("\n" + "="*60)
    print("🚀 EXCEL SYNC AGENT API SERVICE")
    print("="*60)
    print(f"📡 Server: http://{args.host}:{args.port}")
    print(f"💚 Health check: http://{args.host}:{args.port}/health")
    print(f"📊 Status: http://{args.host}:{args.port}/api/agent/status")
    print("="*60 + "\n")
    
    uvicorn.run(
        app,
        host=args.host,
        port=args.port,
        log_level="info"
    )

if __name__ == '__main__':
    main()
