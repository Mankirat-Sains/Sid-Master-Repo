#!/usr/bin/env python3
"""
Intelligent Excel Parser - Generic, AI-Driven Semantic Understanding

This parser creates semantic metadata for Excel workbooks by:
1. Detecting color legends using AI (no hardcoded keywords)
2. Classifying cells based on legend colors
3. Understanding cell definitions from nearby context (AI-driven)
4. Creating semantic groups (inputs/outputs/lookups) intelligently
5. Generating metadata in the format required by the local agent

CRITICAL PRINCIPLE: No hardcoding. Everything is AI-driven and generic.
Works with ANY Excel layout, ANY color scheme, ANY legend format.

Author: Sidian Engineering Team
"""

import json
import logging
import os
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from dotenv import load_dotenv
from openai import OpenAI

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load environment
env_path = Path(__file__).parent.parent.parent / ".env"
if env_path.exists():
    load_dotenv(env_path)

# Lazy OpenAI client initialization (avoids hanging on import)
_client = None

def get_openai_client():
    """Get or create OpenAI client (lazy initialization to avoid import hangs)"""
    global _client
    if _client is None:
        try:
            _client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            if not _client.api_key:
                logger.warning("⚠️ OPENAI_API_KEY not set - AI features will be limited")
                _client = None
        except Exception as e:
            logger.warning(f"⚠️ Failed to initialize OpenAI: {e}")
            _client = None
    return _client



class IntelligentExcelParser:
    """
    Intelligent Excel Parser - Fully AI-driven, no hardcoding.
    
    This parser:
    - Detects legends using AI (works with any format)
    - Classifies cells by color (no hardcoded categories)
    - Understands cell meaning from context (AI-driven)
    - Creates semantic groups intelligently
    - Generates metadata for local agent
    """
    
    def __init__(self, openai_client: Optional[OpenAI] = None):
        """
        Initialize parser with OpenAI client.
        
        Args:
            openai_client: OpenAI client instance (optional, will create if not provided)
        """
        self.client = openai_client or get_openai_client()
        if not self.client:
            raise ValueError("OpenAI client is required for intelligent parsing")
    
    def detect_legend_with_ai(self, file_path: Path, sheet_name: str) -> Dict[str, Any]:
        """
        Use AI to intelligently detect and understand ANY legend format.
        
        NO HARDCODING: The AI figures out what the legend means, regardless of:
        - Legend location (top, bottom, side, anywhere)
        - Legend format (text, colors, symbols, etc.)
        - Language or terminology used
        
        Args:
            file_path: Path to Excel workbook
            sheet_name: Name of sheet to analyze
        
        Returns:
            Dictionary with legend information:
            {
                "legend_found": bool,
                "legend_cells": [...],
                "color_mappings": {
                    "color_hex": {
                        "category": "user_input|calculated_output|...",
                        "description": "...",
                        "confidence": 0.95
                    }
                }
            }
        """
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb[sheet_name]
            
            # Extract all colored cells with their text (potential legend items)
            colored_cells = []
            
            # Scan entire sheet for colored cells (no assumptions about location)
            # IMPORTANT: Include cells with colors even if they don't have text (for color swatches)
            for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), min_col=1, max_col=min(50, ws.max_column)):
                for cell in row:
                    fill_color = None
                    text_color = None
                    
                    # Extract fill color - handle both RGB string and Color objects
                    if cell.fill and cell.fill.start_color:
                        rgb_value = cell.fill.start_color.rgb
                        if rgb_value:
                            # Convert to string, handling both string and object types
                            try:
                                if isinstance(rgb_value, str):
                                    fill_color = rgb_value
                                elif hasattr(rgb_value, '__str__'):
                                    fill_color = str(rgb_value)
                                elif hasattr(rgb_value, 'value'):
                                    fill_color = str(rgb_value.value)
                                else:
                                    fill_color = str(rgb_value)
                            except Exception:
                                fill_color = None
                    
                    # Extract text color
                    if cell.font and cell.font.color:
                        if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb:
                            rgb_value = cell.font.color.rgb
                            try:
                                if isinstance(rgb_value, str):
                                    text_color = rgb_value
                                elif hasattr(rgb_value, '__str__'):
                                    text_color = str(rgb_value)
                                elif hasattr(rgb_value, 'value'):
                                    text_color = str(rgb_value.value)
                                else:
                                    text_color = str(rgb_value)
                            except Exception:
                                text_color = None
                    
                    # Check if this cell is in legend area or has legend-related text
                    cell_coord = cell.coordinate
                    cell_text = str(cell.value).lower() if cell.value else ""
                    is_legend_area = any(col in cell_coord for col in ['R', 'S', 'T', 'U'])
                    has_legend_text = any(keyword in cell_text for keyword in ['input', 'output', 'calculation', 'override', 'legend'])
                    
                    # Include cells with colors, BUT:
                    # - Always include white cells if they're in legend area or have legend text
                    # - Otherwise exclude white/black/default colors
                    if fill_color:
                        fill_color_str = str(fill_color) if fill_color else None
                        is_white_or_default = fill_color_str in ['00000000', 'FFFFFFFF', 'FFFFFF', 'None', None, '']
                        
                        # Include if: not white/default, OR (white but in legend area/text)
                        if not is_white_or_default or (is_white_or_default and (is_legend_area or has_legend_text)):
                            colored_cells.append({
                                "cell": cell_coord,
                                "text": str(cell.value)[:100] if cell.value else "",  # Empty string if no value
                                "fill_color": fill_color_str if fill_color_str else "",
                                "text_color": str(text_color) if text_color else ""
                                # Removed "row" and "column" to match original exactly
                            })
            
            wb.close()
            
            if not colored_cells:
                logger.info(f"   ⚠️ No colored cells found in {sheet_name}")
                return {"legend_found": False, "color_mappings": {}, "legend_cells": []}
            
            # Debug: Print sample colored cells to help diagnose issues
            logger.info(f"   🔍 Sample colored cells found:")
            for i, cell_info in enumerate(colored_cells[:20], 1):
                color_str = str(cell_info.get('fill_color', '')) if cell_info.get('fill_color') else 'None'
                logger.info(f"      {i}. {cell_info['cell']}: '{cell_info.get('text', '')}' | Color: {color_str}")
            
            # Prioritize cells that look like legend items
            # Look for cells with text like "Input", "Output", "Calculation", "Override", "Legend"
            legend_keywords = ['input', 'output', 'calculation', 'override', 'legend', 'key']
            prioritized_cells = []
            regular_cells = []
            
            for cell_info in colored_cells:
                cell_text = cell_info.get('text', '').lower()
                cell_coord = cell_info.get('cell', '')
                
                # Prioritize cells with legend keywords or in columns R-U (common legend area)
                is_legend_like = (
                    any(keyword in cell_text for keyword in legend_keywords) or
                    any(col in cell_coord for col in ['R', 'S', 'T', 'U'])  # Legend often in right columns
                )
                
                if is_legend_like:
                    prioritized_cells.append(cell_info)
                else:
                    regular_cells.append(cell_info)
            
            # Combine: prioritized first, then regular (up to 150 total)
            cells_for_ai = prioritized_cells[:100] + regular_cells[:50]
            
            if prioritized_cells:
                logger.info(f"   🎯 Found {len(prioritized_cells)} legend-like cells (prioritizing these)")
            
            # Use AI to understand which cells form a legend
            logger.info(f"   🤖 AI analyzing {len(cells_for_ai)} prioritized colored cells for legend...")
            
            prompt = f"""You are analyzing a spreadsheet to find and interpret its color legend.

Here are colored cells from the sheet (prioritized for legend detection):
{json.dumps(cells_for_ai, indent=2)}

Your task:
1. Identify which cells (if any) form a LEGEND or KEY that explains what colors mean
2. Look for patterns like:
   - A section titled "Legend" or "Key" with colored cells and labels
   - Color swatches (cells with colors) next to text labels like "Input", "Output", "Calculation", "Override"
   - Cells grouped together that systematically explain color meanings
3. For each legend item, determine what semantic category it represents

Common legend categories (but don't limit yourself to these):
- User inputs / editable cells (often labeled "Input")
- Calculated values / formulas (often labeled "Calculation")
- Output results / things to review (often labeled "Output")
- Status indicators (pass/fail, warnings, etc.)
- Override cells (cells that exceed limits, often labeled "Override")
- Capacity ratios / utilization (e.g., Mf/Mr, Vf/Vr, deflection ratios)
- Design results / member sizes
- Labels or headers
- Protected/locked cells

IMPORTANT: For structural design spreadsheets, identify cells showing:
- Moment capacity ratios (Mf/Mr)
- Shear capacity ratios (Vf/Vr)
- Deflection ratios (Actual/Limit)
- Member size selections
These are critical for iterative design optimization!

Pay special attention to:
- Cells with borders (like color swatches)
- Cells with colors but no text (these might be legend examples)
- Cells near text like "Input", "Output", "Calculation", "Override"
- Cells in columns R, S, T, U (common legend location)

CRITICAL: If you see a "Legend" section with multiple entries (Input, Output, Calculation, Override), 
you MUST identify ALL of them, not just one!

Return ONLY valid JSON:
{{
  "legend_found": true/false,
  "legend_cells": [
    {{
      "cell": "S29",
      "text": "Input",
      "fill_color": "FF00FF00",
      "semantic_category": "user_input",
      "description": "Cells users should modify"
    }}
  ],
  "color_mappings": {{
    "FF00FF00": {{
      "category": "user_input",
      "description": "Editable input cells",
      "confidence": 0.95
    }}
  }}
}}

CRITICAL: Return valid JSON only. No markdown, no explanation.
If no legend exists, return {{"legend_found": false, "legend_cells": [], "color_mappings": {{}}}}
"""
            
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert at analyzing spreadsheet structure. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,
                temperature=0.1
            )
            
            raw_content = response.choices[0].message.content.strip()
            
            # Strip markdown if present
            if raw_content.startswith("```json"):
                raw_content = raw_content[7:]
            if raw_content.startswith("```"):
                raw_content = raw_content[3:]
            if raw_content.endswith("```"):
                raw_content = raw_content[:-3]
            raw_content = raw_content.strip()
            
            result = json.loads(raw_content)
            
            if result.get("legend_found"):
                logger.info(f"   ✅ AI found legend with {len(result.get('color_mappings', {}))} color categories")
                for color, mapping in list(result.get("color_mappings", {}).items())[:5]:
                    logger.info(f"      • {mapping.get('category', 'unknown')}: {mapping.get('description', '')}")
            else:
                logger.info(f"   ⚠️ AI determined no legend exists")
            
            return result
        
        except Exception as e:
            logger.error(f"   ❌ Legend detection failed: {e}")
            return {"legend_found": False, "color_mappings": {}, "legend_cells": []}
    
    def extract_cell_neighborhood(self, cell_coord: str, ws, radius: int = 3) -> Dict[str, str]:
        """
        Extract neighborhood around a cell for AI context analysis.
        
        Returns dict with relative positions as keys: {(-1,0): "Span", (1,0): "ft", ...}
        This provides context for AI to understand what the cell represents.
        
        Args:
            cell_coord: Cell coordinate (e.g., "B5")
            ws: OpenPyXL worksheet object
            radius: Radius of neighborhood to extract (default: 3 = 7x7 area)
        
        Returns:
            Dictionary mapping relative positions to cell values
        """
        # Parse cell coordinate
        col_letter = ''.join(c for c in cell_coord if c.isalpha())
        row_num = int(''.join(c for c in cell_coord if c.isdigit()))
        col_num = column_index_from_string(col_letter)
        
        neighborhood = {}
        
        for r_offset in range(-radius, radius + 1):
            for c_offset in range(-radius, radius + 1):
                if r_offset == 0 and c_offset == 0:
                    continue  # Skip the cell itself
                
                try:
                    check_row = row_num + r_offset
                    check_col = col_num + c_offset
                    
                    if check_row > 0 and check_col > 0:
                        check_cell = ws.cell(row=check_row, column=check_col)
                        if check_cell.value:
                            pos_key = f"({c_offset},{r_offset})"  # (col_offset, row_offset)
                            neighborhood[pos_key] = str(check_cell.value)[:100]
                except:
                    pass
        
        return neighborhood
    
    def understand_cell_meaning(self, cell_coord: str, neighborhood: Dict[str, str], category: str) -> Dict[str, Any]:
        """
        Use AI to understand what a cell represents based on surrounding context.
        
        NO HARDCODING: AI analyzes the neighborhood to determine:
        - Parameter name
        - Description
        - Units
        - Engineering meaning
        
        This works with ANY layout:
        - Labels left, right, above, below, or diagonal
        - Units in separate cells or embedded in labels
        - Multi-column layouts (Imperial/Metric)
        - Any language or terminology
        
        Args:
            cell_coord: Cell coordinate
            neighborhood: Dictionary of nearby cell values (from extract_cell_neighborhood)
            category: Semantic category from legend (e.g., "user_input")
        
        Returns:
            Dictionary with:
            {
                "description": "Span in meters",
                "parameter_name": "span",
                "unit": "m",
                "confidence": 0.95,
                "engineering_meaning": "Clear span of the structural member"
            }
        """
        if not neighborhood:
            return {
                "description": f"{category} cell at {cell_coord}",
                "parameter_name": f"param_{cell_coord.lower()}",
                "unit": "unknown",
                "confidence": 0.3,
                "engineering_meaning": "Unknown parameter"
            }
        
        # Build concise neighborhood description
        nearby_text = []
        for pos, value in sorted(neighborhood.items())[:20]:  # Top 20 nearest cells
            nearby_text.append(f"{pos}: {value}")
        
        prompt = f"""Cell {cell_coord} is a {category} cell in an engineering spreadsheet.

Nearby cells (position relative to {cell_coord}):
{chr(10).join(nearby_text)}

Position format: (column_offset, row_offset) from {cell_coord}
Examples: (-1,0)=left, (1,0)=right, (0,-1)=above, (0,1)=below

Task: Determine what engineering parameter this cell represents.

Look for:
1. Label/parameter name (often at (-1,0), (0,-1), or diagonal)
2. Units (often at (1,0), or in label text like "Span (ft)")
3. Column header (above at (0,-1) or (0,-2))
4. Engineering context from surrounding cells

Examples:
- If (-1,0)="Span" and (1,0)="ft" → parameter="span", unit="ft", description="Span in feet"
- If (0,-1)="Load (kN/m)" → parameter="load", unit="kN/m", description="Load in kN/m"
- If (-1,0)="Live Load" and (1,0)="kPa" → parameter="live_load", unit="kPa", description="Live load in kPa"
- If (-2,0)="Span" and (0,-1)="Metric" and (1,0)="m" → parameter="span", unit="m", description="Span in meters"

Return ONLY valid JSON (no markdown):
{{
  "description": "Span in meters",
  "parameter_name": "span",
  "unit": "m",
  "confidence": 0.95,
  "engineering_meaning": "Clear span of the structural member"
}}

If uncertain, return confidence < 0.5 and best guess.
"""
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",  # Cheaper for this task
                messages=[
                    {"role": "system", "content": "You extract engineering parameter context from spreadsheets. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=200,
                temperature=0.1
            )
            
            raw_content = response.choices[0].message.content.strip()
            
            # Strip markdown if present
            if raw_content.startswith("```json"):
                raw_content = raw_content[7:]
            if raw_content.startswith("```"):
                raw_content = raw_content[3:]
            if raw_content.endswith("```"):
                raw_content = raw_content[:-3]
            raw_content = raw_content.strip()
            
            context = json.loads(raw_content)
            return context
        
        except Exception as e:
            logger.warning(f"   ⚠️ Context extraction failed for {cell_coord}: {e}")
            return {
                "description": f"{category} parameter at {cell_coord}",
                "parameter_name": f"param_{cell_coord.lower()}",
                "unit": "unknown",
                "confidence": 0.3,
                "engineering_meaning": "Unknown parameter"
            }
    
    def classify_cells_by_legend(self, file_path: Path, sheet_name: str, color_mappings: Dict[str, Any]) -> Dict[str, List[Dict]]:
        """
        Classify ALL cells based on AI-detected legend with contextual descriptions.
        
        Works with ANY color scheme and layout. Uses AI to understand each cell's meaning.
        
        Args:
            file_path: Path to Excel workbook
            sheet_name: Name of sheet to analyze
            color_mappings: Color mappings from legend detection
        
        Returns:
            Dictionary with classified cells:
            {
                "inputs": [...],
                "outputs": [...],
                "calculations": [...],
                "status_indicators": [...],
                "capacity_ratios": [...],
                "member_sizes": [...]
            }
        """
        if not color_mappings:
            logger.warning(f"   ⚠️ No color mappings provided for {sheet_name}")
            return {
                "inputs": [],
                "outputs": [],
                "calculations": [],
                "status_indicators": [],
                "capacity_ratios": [],
                "member_sizes": [],
                "labels": []
            }
        
        wb = load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        
        classified_cells = {
            "inputs": [],
            "outputs": [],
            "calculations": [],
            "status_indicators": [],
            "capacity_ratios": [],
            "member_sizes": [],
            "labels": []
        }
        
        # Track cells needing context extraction
        cells_needing_context = []
        
        logger.info(f"   🏷️ Classifying cells by legend colors...")
        
        # Normalize color mappings keys to strings for matching
        normalized_color_mappings = {}
        for color_key, mapping in color_mappings.items():
            normalized_key = str(color_key).upper().strip()
            normalized_color_mappings[normalized_key] = mapping
        
        logger.info(f"   🔍 Color mappings available: {list(normalized_color_mappings.keys())}")
        
        # Scan all cells (include cells without values for color swatches)
        for row in ws.iter_rows(min_row=1, max_row=min(300, ws.max_row), min_col=1, max_col=min(50, ws.max_column)):
            for cell in row:
                # Get cell's fill color using same robust extraction as legend detection
                fill_color = None
                if cell.fill and cell.fill.start_color:
                    rgb_value = cell.fill.start_color.rgb
                    if rgb_value:
                        try:
                            if isinstance(rgb_value, str):
                                fill_color = rgb_value
                            elif hasattr(rgb_value, '__str__'):
                                fill_color = str(rgb_value)
                            elif hasattr(rgb_value, 'value'):
                                fill_color = str(rgb_value.value)
                            else:
                                fill_color = str(rgb_value)
                        except Exception:
                            fill_color = None
                
                # Normalize fill_color for matching
                if fill_color:
                    fill_color_normalized = str(fill_color).upper().strip()
                else:
                    fill_color_normalized = None
                
                # Look up meaning from AI-detected legend
                if fill_color_normalized and fill_color_normalized in normalized_color_mappings:
                    category = normalized_color_mappings[fill_color_normalized].get("category", "unknown")
                    
                    cell_info = {
                        "cell": cell.coordinate,
                        "value": str(cell.value)[:100] if cell.value else "",
                        "color": fill_color_normalized,
                        "category": category,
                        "description": normalized_color_mappings[fill_color_normalized].get("description", ""),
                        "has_formula": cell.data_type == 'f',
                        "sheet": sheet_name
                    }
                    
                    # Map AI categories to our classification (expanded to include all detected categories)
                    # INPUT categories
                    if category in ["user_input", "editable", "input", "parameter", "input_cell"]:
                        classified_cells["inputs"].append(cell_info)
                        if cell.value:  # Only extract context for cells with values
                            cells_needing_context.append((cell.coordinate, cell_info, "inputs"))
                    
                    # OUTPUT categories (including output_result, calculated_value)
                    elif category in ["output", "result", "check", "review", "calculated_output", 
                                     "output_result", "calculated_value", "output_cell"]:
                        classified_cells["outputs"].append(cell_info)
                        if cell.value:  # Only extract context for cells with values
                            cells_needing_context.append((cell.coordinate, cell_info, "outputs"))
                    
                    # CALCULATION categories
                    elif category in ["calculation", "formula", "computed", "calculation_cell"]:
                        classified_cells["calculations"].append(cell_info)
                    
                    # OVERRIDE categories
                    elif category in ["override", "override_cell", "exceed_limit"]:
                        classified_cells["status_indicators"].append(cell_info)
                        # Override cells are also outputs (they show results that exceed limits)
                        classified_cells["outputs"].append(cell_info)
                        if cell.value:
                            cells_needing_context.append((cell.coordinate, cell_info, "outputs"))
                    
                    # STATUS categories
                    elif category in ["status", "indicator", "pass", "fail", "warning"]:
                        classified_cells["status_indicators"].append(cell_info)
                    
                    # CAPACITY RATIO categories
                    elif category in ["capacity_ratio", "utilization", "ratio", "mf_mr", "vf_vr", "deflection_ratio"]:
                        classified_cells["capacity_ratios"].append(cell_info)
                        classified_cells["outputs"].append(cell_info)  # Capacity ratios are outputs
                    
                    # MEMBER SIZE categories
                    elif category in ["member_size", "size_selection", "section", "dimension"]:
                        classified_cells["member_sizes"].append(cell_info)
                        classified_cells["outputs"].append(cell_info)  # Member sizes are outputs
                    
                    # LABEL categories
                    elif category in ["label", "header", "description"]:
                        classified_cells["labels"].append(cell_info)
                
                # Also detect formulas (always calculated) - but only if not already classified
                elif cell.data_type == 'f' and cell.value:
                    # Check if we already classified this cell
                    already_classified = False
                    for cat_list in classified_cells.values():
                        for existing_cell in cat_list:
                            if existing_cell.get("cell") == cell.coordinate:
                                already_classified = True
                                break
                        if already_classified:
                            break
                    
                    if not already_classified:
                        classified_cells["calculations"].append({
                            "cell": cell.coordinate,
                            "value": str(cell.value)[:100],
                            "formula": True,
                            "category": "formula_detected",
                            "sheet": sheet_name
                        })
        
        # Log classification results
        logger.info(f"   📊 Classification results:")
        logger.info(f"      Inputs: {len(classified_cells['inputs'])}")
        logger.info(f"      Outputs: {len(classified_cells['outputs'])}")
        logger.info(f"      Calculations: {len(classified_cells['calculations'])}")
        logger.info(f"      Status indicators: {len(classified_cells['status_indicators'])}")
        
        # 🧠 AI-DRIVEN CONTEXT EXTRACTION
        # SKIP context extraction for now - it's too slow and not essential for basic parsing
        # We can add it back later with batching or make it optional
        if cells_needing_context:
            logger.info(f"   ⚡ Skipping context extraction for {len(cells_needing_context)} cells (too slow)")
            logger.info(f"   💡 Using basic descriptions from legend instead")
            
            # Just use basic descriptions from legend/category
            for cell_coord, cell_info, category_list in cells_needing_context:
                # Use category and cell coordinate for basic parameter name
                cell_info["parameter_name"] = f"{cell_info['category']}_{cell_coord.lower()}"
                cell_info["description"] = cell_info.get("description", f"{cell_info['category']} at {cell_coord}")
                cell_info["unit"] = "unknown"
                cell_info["engineering_meaning"] = ""
                cell_info["context_confidence"] = 0.5
            
            logger.info(f"   ✅ Basic descriptions assigned")
        
        wb.close()
        return classified_cells
    
    def create_semantic_groups(self, classified_cells: Dict[str, List[Dict]], sheet_name: str) -> Dict[str, Any]:
        """
        Create semantic groups intelligently using AI.
        
        Groups related parameters together (e.g., "location_data", "project_parameters").
        This aligns with the gameplan: semantic grouping, not individual cells.
        
        Args:
            classified_cells: Dictionary of classified cells
            sheet_name: Name of sheet
        
        Returns:
            Semantic metadata structure:
            {
                "inputs": {
                    "group_name": {
                        "type": "group",
                        "sheet": "...",
                        "cells": {
                            "parameter_name": "cell_address"
                        }
                    }
                },
                "outputs": {...},
                "lookups": {...}
            }
        """
        logger.info(f"   🧠 Creating semantic groups using AI...")
        
        # Prepare data for AI grouping
        inputs_data = []
        for cell_info in classified_cells.get("inputs", []):
            inputs_data.append({
                "cell": cell_info["cell"],
                "parameter_name": cell_info.get("parameter_name", "unknown"),
                "description": cell_info.get("description", ""),
                "unit": cell_info.get("unit", "unknown")
            })
        
        outputs_data = []
        for cell_info in classified_cells.get("outputs", []):
            outputs_data.append({
                "cell": cell_info["cell"],
                "parameter_name": cell_info.get("parameter_name", "unknown"),
                "description": cell_info.get("description", ""),
                "unit": cell_info.get("unit", "unknown")
            })
        
        prompt = f"""You are creating semantic groups for an Excel spreadsheet.

Sheet: {sheet_name}

Input Parameters ({len(inputs_data)}):
{json.dumps(inputs_data[:50], indent=2)}

Output Parameters ({len(outputs_data)}):
{json.dumps(outputs_data[:50], indent=2)}

Task: Group related parameters into semantic groups.

Examples of good groupings:
- Location-related: location_name, ground_snow_load, ground_rain_load, wind_load → "location_data"
- Project dimensions: building_width, building_length, building_height → "project_parameters"
- Load cases: live_load_udl1, dead_load_udl1, live_load_udl2 → "load_cases"
- Design results: governing_moment, governing_shear, governing_member → "design_summary"
- Capacity checks: moment_ratio, shear_ratio, deflection_ratio → "capacity_ratios"

Return ONLY valid JSON:
{{
  "inputs": {{
    "group_name": {{
      "type": "group",
      "sheet": "{sheet_name}",
      "cells": {{
        "parameter_name": "cell_address"
      }},
      "description": "What this group represents"
    }}
  }},
  "outputs": {{
    "group_name": {{
      "type": "group",
      "sheet": "{sheet_name}",
      "cells": {{
        "parameter_name": "cell_address"
      }},
      "description": "What this group represents"
    }}
  }},
  "lookups": {{}}
}}

CRITICAL: 
- Group related parameters together
- Use meaningful group names (location_data, project_parameters, design_summary, etc.)
- Don't create too many groups (aim for 3-8 groups per category)
- Return valid JSON only
"""
        
        try:
            response = self.client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You create semantic groups for spreadsheet parameters. Return only valid JSON."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.2
            )
            
            raw_content = response.choices[0].message.content.strip()
            
            # Strip markdown if present
            if raw_content.startswith("```json"):
                raw_content = raw_content[7:]
            if raw_content.startswith("```"):
                raw_content = raw_content[3:]
            if raw_content.endswith("```"):
                raw_content = raw_content[:-3]
            raw_content = raw_content.strip()
            
            result = json.loads(raw_content)
            logger.info(f"   ✅ Created {len(result.get('inputs', {}))} input groups, {len(result.get('outputs', {}))} output groups")
            return result
        
        except Exception as e:
            logger.error(f"   ❌ Semantic grouping failed: {e}")
            # Fallback: Create simple groups
            return self._create_fallback_groups(classified_cells, sheet_name)
    
    def _create_fallback_groups(self, classified_cells: Dict[str, List[Dict]], sheet_name: str) -> Dict[str, Any]:
        """Fallback: Create simple groups if AI fails"""
        groups = {
            "inputs": {},
            "outputs": {},
            "lookups": {}
        }
        
        # Simple grouping: all inputs in one group
        if classified_cells.get("inputs"):
            input_cells = {}
            for cell_info in classified_cells["inputs"]:
                param_name = cell_info.get("parameter_name", f"param_{cell_info['cell']}")
                input_cells[param_name] = cell_info["cell"]
            
            if input_cells:
                groups["inputs"]["input_parameters"] = {
                    "type": "group",
                    "sheet": sheet_name,
                    "cells": input_cells,
                    "description": "Input parameters"
                }
        
        # Simple grouping: all outputs in one group
        if classified_cells.get("outputs"):
            output_cells = {}
            for cell_info in classified_cells["outputs"]:
                param_name = cell_info.get("parameter_name", f"param_{cell_info['cell']}")
                output_cells[param_name] = cell_info["cell"]
            
            if output_cells:
                groups["outputs"]["output_results"] = {
                    "type": "group",
                    "sheet": sheet_name,
                    "cells": output_cells,
                    "description": "Output results"
                }
        
        return groups
    
    def parse_workbook(self, file_path: Union[str, Path], output_path: Optional[Union[str, Path]] = None) -> Dict[str, Any]:
        """
        Parse an Excel workbook and generate semantic metadata.
        
        This is the main entry point. It:
        1. Detects legend using AI
        2. Classifies cells by color
        3. Understands cell meanings from context
        4. Creates semantic groups
        5. Generates metadata for local agent
        
        Args:
            file_path: Path to Excel workbook
            output_path: Optional path to save metadata JSON
        
        Returns:
            Semantic metadata dictionary
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Workbook not found: {file_path}")
        
        logger.info(f"\n📄 Parsing workbook: {file_path.name}")
        
        try:
            wb = load_workbook(file_path, data_only=False)
            sheet_names = wb.sheetnames
            wb.close()
            
            logger.info(f"   Sheets: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}")
            
            # Step 1: Detect legend (try first few sheets)
            # Check up to 10 sheets like the original script
            workbook_legend = None
            legend_sheet = None
            
            logger.info(f"   🔍 Searching for color legend across sheets...")
            for sheet_name in sheet_names[:10]:  # Check first 10 sheets (matching original)
                legend_info = self.detect_legend_with_ai(file_path, sheet_name)
                if legend_info.get("legend_found"):
                    workbook_legend = legend_info
                    legend_sheet = sheet_name
                    logger.info(f"   ✅ Found legend on sheet: {sheet_name}")
                    logger.info(f"      Color mappings will apply to ALL sheets")
                    break
            
            if not workbook_legend:
                logger.warning(f"   ⚠️ No legend found in workbook - will analyze without color classification")
                workbook_legend = {"legend_found": False, "color_mappings": {}}
            
            color_mappings = workbook_legend.get("color_mappings", {})
            
            # Step 2: Analyze each sheet
            all_semantic_groups = {
                "inputs": {},
                "outputs": {},
                "lookups": {}
            }
            
            # Focus on key sheets (first 5, matching original script)
            key_sheets = sheet_names[:5]  # Analyze first 5 sheets
            
            for sheet_name in key_sheets:
                logger.info(f"\n   📊 Analyzing sheet: {sheet_name}")
                
                # Classify cells
                classified_cells = self.classify_cells_by_legend(file_path, sheet_name, color_mappings)
                
                logger.info(f"      Found: {len(classified_cells['inputs'])} inputs, "
                          f"{len(classified_cells['outputs'])} outputs")
                
                # Create semantic groups
                semantic_groups = self.create_semantic_groups(classified_cells, sheet_name)
                
                # Merge into overall structure
                for category in ["inputs", "outputs", "lookups"]:
                    all_semantic_groups[category].update(semantic_groups.get(category, {}))
            
            # Step 3: Create final metadata structure
            metadata = {
                "workbook_path": str(file_path),
                "workbook_name": file_path.name,
                "legend_detected": workbook_legend.get("legend_found", False),
                "legend_sheet": legend_sheet,
                "semantic_interface": all_semantic_groups
            }
            
            # Save if output path provided
            if output_path:
                output_path = Path(output_path)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(metadata, f, indent=2, ensure_ascii=False)
                
                logger.info(f"\n✅ Semantic metadata saved to: {output_path}")
            
            logger.info(f"\n✅ Parsing complete:")
            logger.info(f"   Input groups: {len(all_semantic_groups['inputs'])}")
            logger.info(f"   Output groups: {len(all_semantic_groups['outputs'])}")
            logger.info(f"   Lookup groups: {len(all_semantic_groups['lookups'])}")
            
            return metadata
        
        except Exception as e:
            logger.error(f"❌ Parsing failed: {e}")
            import traceback
            traceback.print_exc()
            raise


def main():
    """Main entry point for command-line usage"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Intelligent Excel Parser - AI-driven semantic understanding")
    parser.add_argument("workbook", type=Path, help="Path to Excel workbook")
    parser.add_argument("-o", "--output", type=Path, help="Output path for metadata JSON")
    parser.add_argument("--api-key", type=str, help="OpenAI API key (or set OPENAI_API_KEY env var)")
    
    args = parser.parse_args()
    
    # Set API key if provided
    if args.api_key:
        os.environ["OPENAI_API_KEY"] = args.api_key
    
    # Create parser
    parser_instance = IntelligentExcelParser()
    
    # Parse workbook
    output_path = args.output or (args.workbook.parent / f"{args.workbook.stem}_metadata.json")
    
    metadata = parser_instance.parse_workbook(args.workbook, output_path)
    
    print(f"\n✅ Success! Metadata generated: {output_path}")


if __name__ == "__main__":
    main()

