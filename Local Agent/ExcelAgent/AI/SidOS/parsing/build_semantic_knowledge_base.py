#!/usr/bin/env python3
"""
Build Semantic Knowledge Base for Excel Templates

This creates a deep understanding of spreadsheets by:
1. Analyzing structure (sheets, cells, formulas)
2. Using AI to understand semantic meaning
3. Identifying workflows and dependencies
4. Creating action templates for operations
5. **ENHANCED**: Leveraging Building Code RAG for deep engineering understanding

Based on SheetMind research (https://arxiv.org/pdf/2506.12339)
"""

import json
import sys
from pathlib import Path
from typing import Dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv
from openai import OpenAI
import pandas as pd

# Load environment
# Check if API key is already set in environment (from command line)
api_key_from_env = os.getenv("OPENAI_API_KEY")

# Try multiple possible .env locations (only if key not already set)
if not api_key_from_env:
    possible_env_paths = [
        Path(__file__).parent.parent.parent / ".env",  # trainexcel/.env
        Path(__file__).parent.parent.parent.parent / ".env",  # Visual Studio/.env
        Path(__file__).parent / ".env",  # parsing/.env
    ]
    
    for env_path in possible_env_paths:
        if env_path.exists():
            load_dotenv(env_path, override=False)  # Don't override existing env vars
            break
    else:
        load_dotenv(override=False)  # Try current directory
else:
    # Key already set, just try to load .env for other variables
    possible_env_paths = [
        Path(__file__).parent.parent.parent / ".env",  # trainexcel/.env
    ]
    for env_path in possible_env_paths:
        if env_path.exists():
            load_dotenv(env_path, override=False)
            break

# Lazy client initialization
_client = None
def get_client():
    """Get or create OpenAI client (lazy initialization)"""
    global _client
    if _client is None:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            raise ValueError("OPENAI_API_KEY environment variable must be set. Please set it with: export OPENAI_API_KEY='your-key'")
        _client = OpenAI(api_key=api_key)
    return _client

# Don't create client at import time - wait until it's actually needed
client = None  # Will be created lazily via get_client()

# Import Building Code RAG
try:
    from building_code_rag import get_building_code_rag
    CODE_RAG_AVAILABLE = True
    print("✅ Building Code RAG integration enabled")
except Exception as e:
    CODE_RAG_AVAILABLE = False
    print(f"⚠️ Building Code RAG not available: {e}")


def extract_sheet_sample(file_path: Path, sheet_name: str, max_rows=50, max_cols=20):
    """Extract a sample of sheet data for AI analysis"""
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows, header=None)
        
        # Convert to string representation
        sample_data = []
        for row_idx in range(min(max_rows, len(df))):
            row_data = []
            for col_idx in range(min(max_cols, len(df.columns))):
                cell_value = df.iloc[row_idx, col_idx]
                if pd.notna(cell_value):
                    cell_addr = f"{get_column_letter(col_idx + 1)}{row_idx + 1}"
                    row_data.append(f"{cell_addr}: {str(cell_value)[:50]}")
            if row_data:
                sample_data.append(" | ".join(row_data))
        
        return "\n".join(sample_data[:30])  # First 30 rows
    except Exception as e:
        return f"Error reading sheet: {e}"


def analyze_sheet_semantics(file_name: str, sheet_name: str, sample_data: str, labeled_cells: dict):
    """
    🧠 MULTI-STAGE AI ANALYSIS for intelligent spreadsheet understanding
    
    Stage 1: Understand layout pattern
    Stage 2: Detect sections and groupings  
    Stage 3: Map cells based on learned structure
    
    This approach works for ANY spreadsheet layout without hardcoding!
    """
    
    print(f"      🧠 Multi-stage analysis of '{sheet_name}'...")
    
    # Get building code context
    code_context = ""
    if CODE_RAG_AVAILABLE:
        try:
            code_rag = get_building_code_rag()
            domain_query = f"What type of structural design is this: {file_name} - {sheet_name}?"
            code_results = code_rag.query_all_codes(domain_query, top_k_per_code=1)
            
            if code_results:
                code_context = "\n\n**Building Code Context:**\n"
                for code_name, results in code_results.items():
                    if results:
                        code_context += f"\n[{code_name}]:\n{results[0]['text'][:300]}...\n"
        except Exception as e:
            print(f"         ⚠️ Code RAG failed: {e}")
    
    # ============================================================================
    # STAGE 1: LAYOUT PATTERN UNDERSTANDING
    # ============================================================================
    print(f"      📐 Stage 1: Understanding layout pattern...")
    
    layout_prompt = f"""Analyze the LAYOUT STRUCTURE of this engineering spreadsheet.

Sample Data (first 30 rows):
{sample_data}

Questions:
1. Are there multiple columns for the same parameter (e.g., Imperial/Metric/Use)? Or single columns?
2. How are parameters organized in rows? Are there repeated sections?
3. What row contains column headers? What are they?
4. What column contains parameter labels?
5. Where do users enter values? (which columns?)

Return JSON:
{{
  "column_pattern": "multi-column" or "single-column",
  "column_structure": {{"imperial_column": "C", "metric_column": "E", "use_column": "G"}},
  "header_row": 3,
  "label_column": "B",
  "input_columns": ["C", "E"],
  "repeating_sections": true,
  "section_pattern": "Describe any repeating patterns like UDL 1, UDL 2"
}}
"""
    
    try:
        layout_response = get_client().chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "system", "content": "Analyze spreadsheet layouts. Return only JSON."}, 
                     {"role": "user", "content": layout_prompt}],
            max_tokens=500,
            temperature=0.1
        )
        
        layout_json = layout_response.choices[0].message.content.strip()
        layout_json = layout_json.strip("`").removeprefix("json").strip()
        layout_pattern = json.loads(layout_json)
        print(f"         ✅ Layout: {layout_pattern.get('column_pattern', 'detected')}")
        
    except Exception as e:
        print(f"         ⚠️ Stage 1 failed, using fallback: {e}")
        layout_pattern = {"column_pattern": "unknown"}
    
    # ============================================================================
    # STAGE 2: SECTION DETECTION
    # ============================================================================
    print(f"      📊 Stage 2: Detecting sections...")
    
    sections_prompt = f"""Identify distinct parameter groups/sections in this sheet.

Layout Pattern:
{json.dumps(layout_pattern, indent=2)}

Sample Data:
{sample_data}

Identify sections and their row ranges. Example:
- "UDL 1" (rows 4-9): Span, Trib Width, Live, Dead
- "UDL 2" (rows 10-13): Trib Width, Live, Dead

Return JSON:
{{
  "sections": [
    {{"name": "UDL 1", "row_start": 4, "row_end": 9, "parameters": ["Span", "Trib. Width", "Live", "Dead"]}},
    {{"name": "UDL 2", "row_start": 10, "row_end": 13, "parameters": ["Trib. Width", "Live", "Dead"]}}
  ]
}}
"""
    
    try:
        sections_response = get_client().chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "system", "content": "Identify sections. Return only JSON."},
                     {"role": "user", "content": sections_prompt}],
            max_tokens=600,
            temperature=0.1
        )
        
        sections_json = sections_response.choices[0].message.content.strip()
        sections_json = sections_json.strip("`").removeprefix("json").strip()
        sections = json.loads(sections_json)
        print(f"         ✅ Found {len(sections.get('sections', []))} sections")
        
    except Exception as e:
        print(f"         ⚠️ Stage 2 failed: {e}")
        sections = {"sections": []}
    
    # ============================================================================
    # STAGE 3: CELL MAPPING (Using learned pattern)
    # ============================================================================
    print(f"      🎯 Stage 3: Mapping cells based on pattern...")
    
    col_struct = layout_pattern.get('column_structure', {})
    imp_col = col_struct.get('imperial_column', 'C')
    met_col = col_struct.get('metric_column', 'E')
    
    mapping_prompt = f"""Map each parameter to cells using the pattern you learned.

Layout Pattern:
{json.dumps(layout_pattern, indent=2)}

Sections:
{json.dumps(sections, indent=2)}

Sample Data:
{sample_data}

Labeled Cells:
{json.dumps(labeled_cells, indent=2)}
{code_context}

**CRITICAL RULES**:
1. If multi-column: For row N, Imperial={imp_col}N, Metric={met_col}N (SAME ROW!)
2. If multiple sections have same label (e.g., "Live"), name them: live_load_udl1, live_load_udl2
3. Use actual row numbers from data

Return JSON:
{{
  "purpose": "Calculate structural requirements...",
  "domain": "Beam Design",
  "input_parameters": [
    {{"name": "span", "cell": "C5", "cell_metric": "E5", "label": "Span", "description": "Clear span", "data_type": "float", "unit": "ft or m", "typical_range": [5, 40], "use_cell": "G5"}},
    {{"name": "live_load_udl1", "cell": "C8", "cell_metric": "E8", "label": "Live (UDL 1)", "description": "Live load for first case", "data_type": "float", "unit": "psf or kPa", "typical_range": [0, 10], "use_cell": "G8"}}
  ],
  "output_results": [...],
  "workflow": [...],
  "dependencies": [...]
}}
"""
    
    try:
        response = get_client().chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "system", "content": "You are a structural engineer. Map cells precisely. Return only JSON."},
                     {"role": "user", "content": mapping_prompt}],
            max_tokens=1500,
            temperature=0.2
        )
        
        raw_content = response.choices[0].message.content.strip()
        raw_content = raw_content.strip("`").removeprefix("json").strip()
        
        result = json.loads(raw_content)
        result["layout_pattern"] = layout_pattern
        result["sections"] = sections.get("sections", [])
        
        print(f"      ✅ Multi-stage complete: {len(result.get('input_parameters', []))} params mapped")
        return result
        
    except Exception as e:
        print(f"      ⚠️ Stage 3 failed: {e}")
        return create_fallback_semantics(sheet_name)


def create_fallback_semantics(sheet_name: str):
    """Create basic semantic structure if AI fails"""
    return {
        "purpose": f"Engineering calculations for {sheet_name}",
        "domain": "Structural Engineering",
        "input_parameters": [],
        "output_results": [],
        "workflow": ["Input parameters", "Calculate results", "Review outputs"],
        "dependencies": []
    }


def enhance_parameters_with_code_knowledge(parameters: list, sheet_context: str) -> list:
    """
    **CODE-AWARE ENHANCEMENT**: Enrich parameters with building code understanding
    
    For each parameter, query building codes to understand:
    - What code defines it
    - Code section/clause
    - Typical ranges per code
    - How it's used in code calculations
    - Code requirements/limits
    """
    if not CODE_RAG_AVAILABLE or not parameters:
        return parameters
    
    try:
        code_rag = get_building_code_rag()
        enhanced_params = []
        
        print(f"         🔍 Enhancing {len(parameters[:5])} parameters with code knowledge...")
        
        for param in parameters[:5]:  # Enhance first 5 to save API calls
            param_name = param.get("name", "")
            if not param_name:
                enhanced_params.append(param)
                continue
            
            # Query building codes for this parameter
            code_understanding = code_rag.understand_parameter(param_name, sheet_context)
            
            if code_understanding.get("found"):
                # Add code reference to parameter
                param["code_reference"] = {
                    "code": code_understanding.get("code", ""),
                    "section": code_understanding.get("section", ""),
                    "formula": code_understanding.get("formula", ""),
                    "code_requirement": code_understanding.get("code_requirement", ""),
                    "usage": code_understanding.get("usage", "")
                }
                
                # Update typical_range if code provides better info
                if "typical_range" in code_understanding and code_understanding["typical_range"]:
                    param["typical_range"] = code_understanding["typical_range"]
                
                # Update unit if code provides it
                if "unit" in code_understanding and code_understanding["unit"]:
                    param["unit"] = code_understanding["unit"]
                
                print(f"            ✅ {param_name}: {code_understanding.get('code', 'N/A')}")
            else:
                print(f"            ⚠️ {param_name}: No code reference found")
            
            enhanced_params.append(param)
        
        # Add remaining parameters without enhancement
        if len(parameters) > 5:
            enhanced_params.extend(parameters[5:])
        
        return enhanced_params
        
    except Exception as e:
        print(f"         ⚠️ Parameter enhancement failed: {e}")
        return parameters


def analyze_spreadsheet_workflow(file_name: str, sheets_analysis: dict):
    """Understand the overall workflow across all sheets"""
    
    print(f"   🤖 Analyzing overall spreadsheet workflow...")
    
    # Build summary of all sheets
    sheets_summary = {}
    for sheet_name, sheet_data in sheets_analysis.items():
        if "semantics" in sheet_data:
            sheets_summary[sheet_name] = {
                "purpose": sheet_data["semantics"]["purpose"],
                "domain": sheet_data["semantics"]["domain"],
                "inputs": len(sheet_data["semantics"]["input_parameters"]),
                "outputs": len(sheet_data["semantics"]["output_results"])
            }
    
    prompt = f"""You are analyzing a complete engineering spreadsheet workbook.

File: {file_name}

Sheets Summary:
{json.dumps(sheets_summary, indent=2)}

Provide:

1. **Overall Purpose**: What is this entire workbook for?
2. **Primary Use Case**: What problem does it solve?
3. **User Workflow**: Step-by-step how a user would use this workbook
4. **Action Templates**: Common operations users would perform

Return as JSON:
{{
  "overall_purpose": "...",
  "primary_use_case": "...",
  "user_workflow": [
    {{
      "step": 1,
      "action": "Select building parameters",
      "sheet": "INFO",
      "description": "..."
    }}
  ],
  "action_templates": [
    {{
      "action": "design_beam",
      "description": "Design a new beam member",
      "steps": ["create_member_sheet", "set_span", "set_load", "calculate"],
      "required_inputs": ["span", "load", "material"]
    }}
  ]
}}
"""
    
    try:
        response = get_client().chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are an expert structural engineer analyzing spreadsheet workflows."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1000,
            temperature=0.2
        )
        
        raw_content = response.choices[0].message.content
        
        # Strip markdown code blocks if present
        if raw_content.startswith("```json"):
            raw_content = raw_content[7:]
        if raw_content.startswith("```"):
            raw_content = raw_content[3:]
        if raw_content.endswith("```"):
            raw_content = raw_content[:-3]
        raw_content = raw_content.strip()
        
        result = json.loads(raw_content)
        print(f"   ✅ Workflow analysis complete")
        return result
        
    except Exception as e:
        print(f"   ⚠️ Workflow analysis failed: {e}")
        return {
            "overall_purpose": "Engineering design calculations",
            "primary_use_case": "Structural member design",
            "user_workflow": [],
            "action_templates": []
        }


def get_cell_text_color(cell):
    """Helper to extract RGB text color from openpyxl cell"""
    if cell.font and cell.font.color:
        color_obj = cell.font.color
        # Handle standard RGB colors
        if color_obj.type == 'rgb' and color_obj.rgb:
            return str(color_obj.rgb)
        # Handle Theme Colors (Excel's default palette)
        elif color_obj.type == 'theme':
            # Note: converting theme index to RGB requires a theme map, 
            # but for AI classification, the theme index itself is often enough.
            return f"THEME_{color_obj.theme}"
        # Handle indexed colors
        elif hasattr(color_obj, 'index') and color_obj.index:
            return f"INDEX_{color_obj.index}"
    return None  # Default/auto color


def detect_legend_with_ai(file_path: Path, sheet_name: str) -> dict:
    """
    Use AI to intelligently detect and understand ANY legend format.
    
    No hardcoded keywords, locations, or colors!
    The LLM figures out what the legend means.
    
    ENHANCED: Now extracts and uses TEXT COLOR as a classification signal.
    """
    try:
        wb = load_workbook(file_path, data_only=False)
        ws = wb[sheet_name]
        
        # Extract all colored cells with their text (potential legend items)
        colored_cells = []
        
        for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=50):
            for cell in row:
                if cell.value:
                    fill_color = None
                    text_color = None
                    
                    # Extract fill color
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                        fill_color = str(cell.fill.start_color.rgb) if cell.fill.start_color.rgb else None
                    
                    # Extract text color using helper function
                    text_color = get_cell_text_color(cell)
                    
                    # Include cells with interesting fill colors OR text colors
                    has_interesting_fill = fill_color and fill_color not in ['00000000', 'FFFFFFFF', 'None', None]
                    has_interesting_text = text_color and text_color not in ['00000000', 'FFFFFFFF', 'None', None, 'THEME_1', 'INDEX_1']  # Exclude default black
                    
                    if has_interesting_fill or has_interesting_text:
                        colored_cells.append({
                            "cell": cell.coordinate,
                            "text": str(cell.value)[:100],
                            "fill_color": fill_color,
                            "text_color": text_color
                        })
        
        wb.close()
        
        if not colored_cells:
            print(f"      ⚠️ No colored cells found")
            return {"legend_found": False, "color_mappings": {}}
        
        # Use AI to understand which cells form a legend
        print(f"      🤖 AI analyzing {len(colored_cells)} colored cells for legend...")
        
        prompt = f"""You are analyzing a spreadsheet to find and interpret its color legend.
I have provided both the FILL (Background) and TEXT (Font) colors for each cell.

Here are ALL the colored cells in the sheet:
{json.dumps(colored_cells[:50], indent=2)}

Your task:
1. Identify which cells (if any) form a LEGEND or KEY that explains what colors mean
2. For each legend item, determine what semantic category it represents
3. **CRITICAL**: Pay attention to TEXT COLORS as well as fill colors:
   - Red text (FFFF0000 or similar) often indicates errors, failures, or overridden values
   - Black text (00000000 or default) is typically normal/calculated values
   - Text color can be the PRIMARY signal even if fill color is white/default
   - In engineering sheets, red text often means "Override" or "Fail" status

Common legend categories (but don't limit yourself to these):
- User inputs / editable cells (often light gray fill, black or red text)
- Calculated values / formulas (often white fill, black text)
- Output results / things to review (often light gray fill, black text)
- Status indicators (pass/fail, warnings, etc.) - often red text for failures
- Override cells (values that exceed limits) - often red text
- Capacity ratios / utilization (e.g., Mf/Mr, Vf/Vr, deflection ratios)
- Design results / member sizes
- Labels or headers
- Protected/locked cells

IMPORTANT: For structural design spreadsheets, identify cells showing:
- Moment capacity ratios (Mf/Mr)
- Shear capacity ratios (Vf/Vr)
- Deflection ratios (Actual/Limit)
- Member size selections
These are critical for iterative design optimization!

Return ONLY valid JSON:
{{
  "legend_found": true/false,
  "legend_cells": [
    {{
      "cell": "M1",
      "text": "INPUT",
      "fill_color": "FFF4B084",
      "text_color": "00000000",
      "semantic_category": "user_input",
      "description": "Cells users should modify"
    }}
  ],
  "color_mappings": {{
    "FILL_FFF4B084_TEXT_00000000": {{
      "category": "user_input",
      "description": "Light gray background with black text - editable input cells",
      "confidence": 0.95
    }},
    "TEXT_FFFF0000": {{
      "category": "override",
      "description": "Red text indicating a value that has been manually overridden or exceeds limits",
      "confidence": 0.9
    }},
    "FILL_F2F2F2_TEXT_FFFF0000": {{
      "category": "user_input",
      "description": "Light gray background with red text - input cells that may need attention",
      "confidence": 0.85
    }}
  }}
}}

CRITICAL: 
- Create mappings that combine Fill + Text colors when they work together
- Also create mappings for TEXT color alone when it's the primary signal
- Use format: "FILL_{fill_color}_TEXT_{text_color}" for combined, "TEXT_{text_color}" for text-only
- Return valid JSON only. No markdown, no explanation.
If no legend exists, return {{"legend_found": false, "legend_cells": [], "color_mappings": {{}}}}
"""
        
        response = get_client().chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are an expert at analyzing spreadsheet structure. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=1500,
            temperature=0.1
        )
        
        raw_content = response.choices[0].message.content.strip()
        
        # Strip markdown if present
        if raw_content.startswith("```json"):
            raw_content = raw_content[7:]
        if raw_content.startswith("```"):
            raw_content = raw_content[3:]
        if raw_content.endswith("```"):
            raw_content = raw_content[:-3]
        raw_content = raw_content.strip()
        
        result = json.loads(raw_content)
        
        if result.get("legend_found"):
            print(f"      ✅ AI found legend with {len(result.get('color_mappings', {}))} color categories")
            for color, mapping in result.get("color_mappings", {}).items():
                print(f"         • {mapping.get('category', 'unknown')}: {mapping.get('description', '')}")
        else:
            print(f"      ⚠️ AI determined no legend exists")
        
        return result
        
    except Exception as e:
        print(f"      ❌ Legend detection failed: {e}")
        return {"legend_found": False, "color_mappings": {}}


def extract_cell_neighborhood(cell_coord: str, ws, radius: int = 3) -> Dict[str, str]:
    """
    Extract 7x7 neighborhood around a cell for AI context analysis
    
    Returns dict with relative positions as keys: {(-1,0): "Span", (1,0): "ft", ...}
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    
    # Parse cell coordinate
    col_letter = ''.join(c for c in cell_coord if c.isalpha())
    row_num = int(''.join(c for c in cell_coord if c.isdigit()))
    col_num = column_index_from_string(col_letter)
    
    neighborhood = {}
    
    for r_offset in range(-radius, radius + 1):
        for c_offset in range(-radius, radius + 1):
            if r_offset == 0 and c_offset == 0:
                continue  # Skip the cell itself
            
            try:
                check_row = row_num + r_offset
                check_col = col_num + c_offset
                
                if check_row > 0 and check_col > 0:
                    check_cell = ws.cell(row=check_row, column=check_col)
                    if check_cell.value:
                        pos_key = f"({c_offset},{r_offset})"  # (col_offset, row_offset)
                        neighborhood[pos_key] = str(check_cell.value)[:50]
            except:
                pass
    
    return neighborhood


def get_contextual_cell_description(cell_coord: str, neighborhood: Dict, category: str) -> Dict:
    """
    Use AI to understand what this specific cell represents based on surrounding context
    
    This handles ANY layout:
    - Labels left, right, above, below, or diagonal
    - Units in separate cells or embedded in labels
    - Multi-column layouts (Imperial/Metric)
    """
    
    if not neighborhood:
        return {
            "description": f"{category} cell at {cell_coord}",
            "parameter_name": "unknown",
            "unit": "unknown",
            "confidence": 0.3
        }
    
    # Build concise neighborhood description
    nearby_text = []
    for pos, value in sorted(neighborhood.items())[:15]:  # Top 15 nearest cells
        nearby_text.append(f"{pos}: {value}")
    
    prompt = f"""Cell {cell_coord} is an INPUT cell in an engineering spreadsheet.

Nearby cells (position relative to {cell_coord}):
{chr(10).join(nearby_text)}

Position format: (column_offset, row_offset) from {cell_coord}
Examples: (-1,0)=left, (1,0)=right, (0,-1)=above, (0,1)=below

Task: Determine what engineering parameter this input cell represents.

Look for:
1. Label/parameter name (often at (-1,0), (0,-1), or diagonal)
2. Units (often at (1,0), or in label text like "Span (ft)")
3. Column header (above at (0,-1) or (0,-2))

Examples:
- If (-1,0)="Span" and (1,0)="ft" → parameter="span", unit="ft", description="Span in feet"
- If (0,-1)="Load (kN/m)" → parameter="load", unit="kN/m", description="Load in kN/m"
- If (-1,0)="Live Load" and (1,0)="kPa" → parameter="live_load", unit="kPa", description="Live load in kPa"
- If (-2,0)="Span" and (0,-1)="Metric" and (1,0)="m" → parameter="span", unit="m", description="Span in meters"

Return ONLY valid JSON (no markdown):
{{
  "description": "Span in feet",
  "parameter_name": "span",
  "unit": "ft",
  "confidence": 0.95
}}

If uncertain, return confidence < 0.5 and best guess.
"""
    
    try:
        response = get_client().chat.completions.create(
            model="gpt-4o-mini",  # Cheaper for this task
            messages=[
                {"role": "system", "content": "You extract engineering parameter context from spreadsheets. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=150,
            temperature=0.1
        )
        
        raw_content = response.choices[0].message.content.strip()
        
        # Strip markdown if present
        if raw_content.startswith("```json"):
            raw_content = raw_content[7:]
        if raw_content.startswith("```"):
            raw_content = raw_content[3:]
        if raw_content.endswith("```"):
            raw_content = raw_content[:-3]
        raw_content = raw_content.strip()
        
        context = json.loads(raw_content)
        return context
        
    except Exception as e:
        print(f"         ⚠️ Context extraction failed for {cell_coord}: {e}")
        return {
            "description": f"Input parameter at {cell_coord}",
            "parameter_name": "unknown",
            "unit": "unknown",
            "confidence": 0.3
        }


def classify_cells_by_ai_legend(file_path: Path, sheet_name: str, color_mappings: dict) -> dict:
    """
    Classify ALL cells based on AI-detected legend with contextual descriptions.
    Works with any color scheme and layout!
    """
    if not color_mappings:
        return {"inputs": [], "outputs": [], "calculations": [], "status_indicators": []}
    
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    classified_cells = {
        "inputs": [],
        "outputs": [],
        "calculations": [],
        "status_indicators": [],
        "capacity_ratios": [],  # Critical for iterative design!
        "member_sizes": [],
        "labels": []
    }
    
    # Track cells to process for context extraction
    cells_needing_context = []
    
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=50):
        for cell in row:
            if not cell.value:
                continue
            
            # Get cell's fill color
            fill_color = None
            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                fill_color = str(cell.fill.start_color.rgb)
            
            # Get cell's text color
            text_color = get_cell_text_color(cell)
            
            # Look up meaning from AI-detected legend
            # Try combined FILL_TEXT mapping first, then TEXT-only, then FILL-only
            category = None
            mapping_key = None
            mapping_desc = ""
            
            # Try combined mapping: FILL_{fill}_TEXT_{text}
            if fill_color and text_color:
                combined_key = f"FILL_{fill_color}_TEXT_{text_color}"
                if combined_key in color_mappings:
                    category = color_mappings[combined_key].get("category", "unknown")
                    mapping_key = combined_key
                    mapping_desc = color_mappings[combined_key].get("description", "")
            
            # Try text-only mapping: TEXT_{text}
            if not category and text_color:
                text_key = f"TEXT_{text_color}"
                if text_key in color_mappings:
                    category = color_mappings[text_key].get("category", "unknown")
                    mapping_key = text_key
                    mapping_desc = color_mappings[text_key].get("description", "")
            
            # Fall back to fill-only mapping
            if not category and fill_color:
                if fill_color in color_mappings:
                    category = color_mappings[fill_color].get("category", "unknown")
                    mapping_key = fill_color
                    mapping_desc = color_mappings[fill_color].get("description", "")
            
            if category:
                cell_info = {
                    "cell": cell.coordinate,
                    "value": str(cell.value)[:100],
                    "fill_color": fill_color,
                    "text_color": text_color,
                    "color": mapping_key,
                    "category": category,
                    "description": mapping_desc,
                    "has_formula": cell.data_type == 'f'
                }
                
                # For INPUT cells, extract contextual description
                if category in ["user_input", "editable", "input", "parameter"]:
                    cells_needing_context.append((cell.coordinate, cell_info, "inputs"))
                    
                # Map AI categories to our classification
                if category in ["user_input", "editable", "input", "parameter"]:
                    classified_cells["inputs"].append(cell_info)
                elif category in ["output", "result", "check", "review"]:
                    classified_cells["outputs"].append(cell_info)
                elif category in ["calculation", "formula", "computed"]:
                    classified_cells["calculations"].append(cell_info)
                elif category in ["status", "indicator", "pass", "fail", "warning"]:
                    classified_cells["status_indicators"].append(cell_info)
                elif category in ["capacity_ratio", "utilization", "ratio", "mf_mr", "vf_vr", "deflection_ratio"]:
                    classified_cells["capacity_ratios"].append(cell_info)
                elif category in ["member_size", "size_selection", "section", "dimension"]:
                    classified_cells["member_sizes"].append(cell_info)
                elif category in ["label", "header", "description"]:
                    classified_cells["labels"].append(cell_info)
            
            # Also detect formulas (always calculated)
            elif cell.data_type == 'f':
                classified_cells["calculations"].append({
                    "cell": cell.coordinate,
                    "value": str(cell.value)[:100],
                    "formula": True,
                    "category": "formula_detected"
                })
    
    # 🧠 AI-DRIVEN CONTEXT EXTRACTION for input cells
    # COST: ~$0.05 per 250 cells with GPT-4o-mini (~400 tokens × $0.15/1M input)
    # TIME: ~2 minutes per sheet (250 calls × 0.5s each)
    if cells_needing_context:
        print(f"      🧠 Extracting context for {len(cells_needing_context)} input cells...")
        print(f"         Estimated cost: ~${len(cells_needing_context) * 0.0002:.3f}")
        
        for i, (cell_coord, cell_info, category_list) in enumerate(cells_needing_context, 1):
            # Extract 7x7 neighborhood
            neighborhood = extract_cell_neighborhood(cell_coord, ws, radius=3)
            
            # Get AI context description
            context = get_contextual_cell_description(
                cell_coord,
                neighborhood,
                cell_info["category"]
            )
            
            # Update cell info with context
            cell_info["description"] = context["description"]
            cell_info["parameter_name"] = context.get("parameter_name", "unknown")
            cell_info["unit"] = context.get("unit", "unknown")
            cell_info["context_confidence"] = context.get("confidence", 0.5)
            
            if i % 10 == 0:
                print(f"         Processed {i}/{len(cells_needing_context)} cells...")
        
        print(f"      ✅ Context extraction complete")
    
    wb.close()
    return classified_cells


def analyze_excel_file_semantic(file_path: Path):
    """Deeply analyze Excel file with AI-driven adaptive color understanding"""
    
    print(f"\n📄 Analyzing: {file_path.name}")
    print(f"   Path: {file_path}")
    
    try:
        wb = load_workbook(file_path, data_only=False)
        
        file_analysis = {
            "file_name": file_path.name,
            "file_path": str(file_path),
            "total_sheets": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "workbook_legend": None,  # Will be populated after legend detection
            "sheets": {}
        }
        
        print(f"   Sheets ({len(wb.sheetnames)}): {', '.join(wb.sheetnames[:5])}{'...' if len(wb.sheetnames) > 5 else ''}")
        
        # ============================================================================
        # STEP 1: DETECT LEGEND ONCE AT WORKBOOK LEVEL (applies to all sheets!)
        # ============================================================================
        print(f"\n   🔍 Searching for color legend across all sheets...")
        
        workbook_legend_info = {"legend_found": False, "color_mappings": {}}
        legend_sheet_name = None
        
        # Try to find legend on any sheet (typically on INFO or first sheet)
        for sheet_name in wb.sheetnames[:10]:  # Check first 10 sheets for legend
            legend_info = detect_legend_with_ai(file_path, sheet_name)
            if legend_info.get("legend_found"):
                workbook_legend_info = legend_info
                legend_sheet_name = sheet_name
                print(f"   ✅ Found workbook legend on sheet: {sheet_name}")
                print(f"      Color mappings will apply to ALL sheets")
                break
        
        if not workbook_legend_info.get("legend_found"):
            print(f"   ⚠️ No legend found in workbook - will analyze without color classification")
        
        # Extract color mappings to apply to all sheets
        workbook_color_mappings = workbook_legend_info.get("color_mappings", {})
        
        # Store workbook-level legend info
        file_analysis["workbook_legend"] = {
            **workbook_legend_info,
            "legend_sheet": legend_sheet_name,
            "applies_to_all_sheets": True
        }
        
        # Analyze each sheet (focus on key sheets)
        key_sheets = wb.sheetnames[:5]  # First 5 sheets
        
        for sheet_name in key_sheets:
            print(f"\n   📊 Sheet: {sheet_name}")
            sheet = wb[sheet_name]
            
            # Use workbook-level color mappings for this sheet
            color_mappings = workbook_color_mappings
            
            # ============================================================================
            # STEP 2: CLASSIFY ALL CELLS BY LEGEND
            # ============================================================================
            classified_cells = classify_cells_by_ai_legend(file_path, sheet_name, color_mappings)
            
            print(f"      ✅ Classified {len(classified_cells['inputs'])} input cells from legend")
            
            # ============================================================================
            # STEP 3: EXTRACT LABELED CELLS (Enhanced with color info)
            # ============================================================================
            labeled_cells = {}
            for row_idx in range(1, min(100, sheet.max_row + 1)):
                for col_idx in range(1, min(10, sheet.max_column + 1)):
                    cell = sheet.cell(row_idx, col_idx)
                    if cell.value and isinstance(cell.value, str):
                        next_cell = sheet.cell(row_idx, col_idx + 1)
                        if next_cell.value is not None:
                            label_lower = str(cell.value).lower()
                            if any(kw in label_lower for kw in ['span', 'load', 'length', 'width', 'height', 
                                                                  'strength', 'grade', 'pressure', 'force']):
                                next_addr = f"{get_column_letter(col_idx + 1)}{row_idx}"
                                
                                # Check if this cell has color classification
                                cell_color_info = None
                                for input_cell in classified_cells.get('inputs', []):
                                    if input_cell['cell'] == next_addr:
                                        cell_color_info = {
                                            "category": input_cell.get('category'),
                                            "color": input_cell.get('color')
                                        }
                                        break
                                
                                labeled_cells[str(cell.value)] = next_addr if not cell_color_info else {
                                    "cell": next_addr,
                                    "color_info": cell_color_info
                                }
            
            print(f"      Found {len(labeled_cells)} labeled cells")
            
            # ============================================================================
            # STEP 4: GET SAMPLE DATA
            # ============================================================================
            sample_data = extract_sheet_sample(file_path, sheet_name)
            
            # ============================================================================
            # STEP 5: AI SEMANTIC ANALYSIS
            # ============================================================================
            semantics = analyze_sheet_semantics(
                file_path.name, 
                sheet_name, 
                sample_data, 
                labeled_cells
            )
            
            # ============================================================================
            # STEP 6: CODE-AWARE ENHANCEMENT
            # ============================================================================
            if "input_parameters" in semantics:
                sheet_context = f"{file_path.name} - {sheet_name}: {semantics.get('purpose', '')}"
                semantics["input_parameters"] = enhance_parameters_with_code_knowledge(
                    semantics["input_parameters"],
                    sheet_context
                )
            
            # ============================================================================
            # STEP 7: STORE EVERYTHING (including color legend)
            # ============================================================================
            file_analysis["sheets"][sheet_name] = {
                "sheet_name": sheet_name,
                "dimensions": f"{sheet.max_row} rows x {sheet.max_column} columns",
                "labeled_cells": labeled_cells,
                "color_legend": {
                    **workbook_legend_info,  # Workbook-level legend
                    "applied_from_sheet": legend_sheet_name,  # Track which sheet had the legend
                    "applied_to_sheet": sheet_name  # This sheet
                },
                "classified_cells": {  # Cells classified by color
                    "inputs": classified_cells.get("inputs", []),
                    "outputs": classified_cells.get("outputs", []),
                    "calculations": classified_cells.get("calculations", []),
                    "capacity_ratios": classified_cells.get("capacity_ratios", []),
                    "member_sizes": classified_cells.get("member_sizes", [])
                },
                "semantics": semantics
            }
        
        wb.close()
        
        # Analyze overall workflow
        workflow_analysis = analyze_spreadsheet_workflow(
            file_path.name,
            file_analysis["sheets"]
        )
        file_analysis["workflow_analysis"] = workflow_analysis
        
        print(f"\n   ✅ Complete analysis finished")
        return file_analysis
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return None


def scan_excel_templates():
    """Scan and deeply analyze Excel Templates"""
    
    # UPDATED: Correct path to Excel Templates
    templates_path = Path("/Users/jameshinsperger/Desktop/Desktop - MacBook Pro (2)/Sidian/Vibeeng/Excel Templates")
    
    if not templates_path.exists():
        print(f"❌ Path not found: {templates_path}")
        print(f"   Please ensure Excel files are in: {templates_path}")
        return {}
    
    print(f"📂 Scanning: {templates_path}")
    
    knowledge_base = {}
    excel_files = list(templates_path.glob("*.xlsx")) + list(templates_path.glob("*.xlsm"))
    
    print(f"📊 Found {len(excel_files)} Excel files\n")
    
    for excel_file in excel_files:
        analysis = analyze_excel_file_semantic(excel_file)
        if analysis:
            # Store with full path included in the analysis
            analysis["file_path"] = str(excel_file.resolve())
            knowledge_base[excel_file.name] = analysis
    
    return knowledge_base


def save_knowledge_base(kb):
    """
    Save semantic knowledge base
    
    **ENHANCED**: Now includes building code references for each parameter
    """
    
    kb_path = Path("excel_knowledge_base.json")
    
    # Add metadata to indicate this is code-enhanced
    kb_metadata = {
        "_metadata": {
            "version": "2.0",
            "enhanced_with_building_codes": CODE_RAG_AVAILABLE,
            "generation_timestamp": str(Path(__file__).stat().st_mtime),
            "description": "Semantic knowledge base enhanced with building code understanding"
        },
        **kb
    }
    
    with open(kb_path, 'w') as f:
        json.dump(kb_metadata, f, indent=2)
    
    print(f"\n✅ CODE-ENHANCED Semantic knowledge base saved: {kb_path}")
    print(f"📊 Total files: {len(kb)}")
    if CODE_RAG_AVAILABLE:
        print(f"🏗️  Enhanced with building code references!")
    
    # Print summary
    print("\n📋 Semantic Knowledge Base Summary:")
    for file_name, info in kb.items():
        print(f"\n  {file_name}:")
        if "workflow_analysis" in info:
            print(f"    Purpose: {info['workflow_analysis'].get('overall_purpose', 'N/A')}")
            print(f"    Use Case: {info['workflow_analysis'].get('primary_use_case', 'N/A')}")
        print(f"    Analyzed Sheets: {len(info['sheets'])}")
        
        # Show code-enhanced parameters
        code_enhanced_count = 0
        if "sheets" in info:
            for sheet_data in info["sheets"].values():
                if "semantics" in sheet_data and "input_parameters" in sheet_data["semantics"]:
                    for param in sheet_data["semantics"]["input_parameters"]:
                        if "code_reference" in param:
                            code_enhanced_count += 1
        
        if code_enhanced_count > 0:
            print(f"    🏗️  Code-Enhanced Parameters: {code_enhanced_count}")
        
        # Show action templates
        if "workflow_analysis" in info and "action_templates" in info["workflow_analysis"]:
            templates = info["workflow_analysis"]["action_templates"]
            if templates:
                print(f"    Action Templates: {len(templates)}")
                for template in templates[:2]:
                    print(f"      - {template.get('action', 'N/A')}: {template.get('description', 'N/A')}")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Build Semantic Excel Knowledge Base")
    parser.add_argument("--file", type=str, help="Path to specific Excel file to analyze (optional)")
    parser.add_argument("--output", type=str, help="Output JSON file path (default: excel_knowledge_base.json)")
    
    args = parser.parse_args()
    
    print("🔧 Building Semantic Excel Knowledge Base...")
    print("Based on SheetMind multi-agent framework research\n")
    
    if args.file:
        # Analyze single file
        file_path = Path(args.file)
        if not file_path.exists():
            print(f"❌ File not found: {file_path}")
            sys.exit(1)
        
        print(f"📄 Analyzing single file: {file_path.name}\n")
        analysis = analyze_excel_file_semantic(file_path)
        
        if analysis:
            analysis["file_path"] = str(file_path.resolve())
            kb = {file_path.name: analysis}
            
            output_path = args.output or "excel_knowledge_base.json"
            kb_metadata = {
                "_metadata": {
                    "version": "2.0",
                    "enhanced_with_building_codes": CODE_RAG_AVAILABLE,
                    "generation_timestamp": str(Path(__file__).stat().st_mtime),
                    "description": "Semantic knowledge base enhanced with building code understanding"
                },
                **kb
            }
            
            with open(output_path, 'w') as f:
                json.dump(kb_metadata, f, indent=2)
            
            print(f"\n✅ Semantic knowledge base saved: {output_path}")
            print(f"📊 File analyzed: {file_path.name}")
        else:
            print("\n❌ Analysis failed!")
            sys.exit(1)
    else:
        # Scan directory
        kb = scan_excel_templates()
        
        if kb:
            output_path = args.output or "excel_knowledge_base.json"
            save_knowledge_base(kb)
        else:
            print("\n❌ No files analyzed!")
            sys.exit(1)

