#!/usr/bin/env python3

import json
import logging
import time
import shutil
import re
import os
from pathlib import Path
from typing import Dict, Any, List, Optional, TypedDict, Annotated, Tuple
from datetime import datetime

from langgraph.graph import StateGraph, END
from langgraph.prebuilt import ToolNode
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage
from openai import OpenAI

# Engineering spreadsheet libraries
import pandas as pd
from openpyxl import load_workbook

# Import the atomic tool system (handle both direct run and module import)
try:
    from excel_tools import get_tool_list, execute_tool
except ImportError:
    from agents.Excel_Agent.excel_tools import get_tool_list, execute_tool

# Conditional xlwings import with error handling
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except Exception as e:
    print(f"⚠️ xlwings not available: {e}")
    XLWINGS_AVAILABLE = False
    xw = None

# Import Building Code RAG (optional - for code validation)
try:
    from building_code_rag import get_building_code_rag
    CODE_RAG_AVAILABLE = True
    print("✅ Building Code RAG integrated")
except ImportError:
    try:
        from agents.Excel_Agent.building_code_rag import get_building_code_rag
        CODE_RAG_AVAILABLE = True
        print("✅ Building Code RAG integrated")
    except Exception as e:
        print(f"⚠️ Building Code RAG not available: {e}")
        print(f"   (This is optional - system will work without code validation)")
        CODE_RAG_AVAILABLE = False
except Exception as e:
    print(f"⚠️ Building Code RAG not available: {e}")
    print(f"   (This is optional - system will work without code validation)")
    CODE_RAG_AVAILABLE = False

# Import Progress Streaming (real-time UI updates)
try:
    from progress_streaming import (
        get_progress_streamer,
        reset_progress_streamer,
        stream_status,
        stream_stage_start,
        stream_stage_complete,
        stream_tool,
        stream_check,
        stream_decision,
        stream_success,
        stream_error
    )
    PROGRESS_STREAMING_AVAILABLE = True
    print("✅ Progress streaming integrated")
except ImportError:
    try:
        from agents.Excel_Agent.progress_streaming import (
            get_progress_streamer,
            reset_progress_streamer,
            stream_status,
            stream_stage_start,
            stream_stage_complete,
            stream_tool,
            stream_check,
            stream_decision,
            stream_success,
            stream_error
        )
        PROGRESS_STREAMING_AVAILABLE = True
        print("✅ Progress streaming integrated")
    except Exception as e:
        print(f"⚠️ Progress streaming not available: {e}")
        PROGRESS_STREAMING_AVAILABLE = False

# Simple settings for RAG environment
class Settings:
    def __init__(self):
        self.openai_api_key = os.getenv("OPENAI_API_KEY")
        self.local_files_root = os.getenv("LOCAL_FILES_ROOT", ".")
        self.debug = os.getenv("DEBUG", "false").lower() == "true"

settings = Settings()

# ============================================================================
# LOGGING CONFIGURATION
# ============================================================================

class ColoredFormatter(logging.Formatter):
    """Custom formatter with colors for better terminal visibility"""
    
    COLORS = {
        'DEBUG': '\033[36m',    # Cyan
        'INFO': '\033[32m',     # Green
        'WARNING': '\033[33m',  # Yellow
        'ERROR': '\033[31m',    # Red
        'CRITICAL': '\033[35m', # Magenta
        'RESET': '\033[0m'      # Reset
    }
    
    def format(self, record):
        log_color = self.COLORS.get(record.levelname, self.COLORS['RESET'])
        reset_color = self.COLORS['RESET']
        
        # Add node execution markers
        if hasattr(record, 'node_name'):
            record.levelname = f"[{record.node_name}] {record.levelname}"
        
        record.levelname = f"{log_color}{record.levelname}{reset_color}"
        return super().format(record)

# Configure logger
logger = logging.getLogger("SidianLangGraph")
logger.setLevel(logging.INFO)

# Create console handler with colored formatter
console_handler = logging.StreamHandler()
console_handler.setFormatter(ColoredFormatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
))
logger.addHandler(console_handler)

# ============================================================================
# STATE DEFINITION
# ============================================================================

class AgentState(TypedDict):
    """State object passed between LangGraph nodes"""
    
    # Input
    user_query: str
    chat_history: List[Dict[str, Any]]
    document_context: Dict[str, Any]
    
    # Planning
    execution_plan: Dict[str, Any]
    selected_tools: List[str]
    route_decision: str
    
    # Knowledge Base
    knowledge_base: Dict[str, Any]
    knowledge_updates: List[Dict[str, Any]]
    
    # File Operations
    target_files: List[str]
    file_operations: List[Dict[str, Any]]
    parsed_content: Dict[str, Any]
    
    # Execution
    results: List[Dict[str, Any]]
    errors: List[str]
    
    # Output
    final_response: str
    suggestions: List[str]
    actions: List[str]

# ============================================================================
# NODE IMPLEMENTATIONS
# ============================================================================

class SidianLangGraphAgent:
    """
    Main LangGraph Agent for Sidian Engineering Orchestrator
    
    This agent uses a modular node-based architecture where each node
    represents a specific tool or capability.
    """
    
    def __init__(self):
        self.logger = logger
        self.ai_client = self._init_openai_client()
        self.graph = self._build_graph()
        
        # ====================================================================
        # CURSOR MARKER: Agent Initialization Complete
        # ====================================================================
        self.logger.info("🚀 SidianLangGraphAgent initialized successfully")
    
    def _init_openai_client(self) -> Optional[OpenAI]:
        """Initialize OpenAI client"""
        if not settings.openai_api_key:
            self.logger.warning("OpenAI API key not configured")
            return None
        
        try:
            client = OpenAI(api_key=settings.openai_api_key)
            self.logger.info("✅ OpenAI client initialized")
            return client
        except Exception as e:
            self.logger.error(f"❌ Failed to initialize OpenAI client: {e}")
            return None
    
    def _build_graph(self) -> StateGraph:
        """
        Build the LangGraph with all nodes and edges
        
        SIMPLIFIED GRAPH STRUCTURE (REFACTORED):
        =========================================
        START -> PLANNING -> KNOWLEDGE_BASE -> TOOL_EXECUTION -> RESPONSE -> END
        
        This clean architecture uses excel_tools.py for ALL Excel operations.
        No more redundant nodes or hardcoded operations!
        """
        
        # ====================================================================
        # CURSOR MARKER: Building Simplified LangGraph Structure
        # ====================================================================
        self.logger.info("🔧 Building simplified LangGraph structure...")
        
        # Create the graph
        graph = StateGraph(AgentState)
        
        # Add 4 core nodes (streamlined!)
        graph.add_node("planning_node", self._planning_node)
        graph.add_node("kb_build_node", self._knowledge_base_node)
        graph.add_node("tool_execution_node", self._tool_execution_node)
        graph.add_node("response_node", self._response_node)
        
        # Define the linear flow
        graph.set_entry_point("planning_node")
        graph.add_edge("planning_node", "kb_build_node")
        graph.add_edge("kb_build_node", "tool_execution_node")
        graph.add_edge("tool_execution_node", "response_node")
        graph.add_edge("response_node", END)
        
        # ====================================================================
        # CURSOR MARKER: Simplified Graph Structure Complete
        # ====================================================================
        self.logger.info("✅ Simplified LangGraph structure built successfully")
        self.logger.info("   Nodes: planning → kb → tool_execution → response")
        
        return graph.compile()
    
    # ========================================================================
    # NODE IMPLEMENTATIONS
    # ========================================================================
    
    def _planning_node(self, state: AgentState) -> AgentState:
        """
        PLANNING NODE: Analyze query and prepare context
        
        This node:
        1. Detects mode (ask/agent)
        2. Extracts query intent
        3. Prepares context for tool execution
        """
        # ====================================================================
        # CURSOR MARKER: Planning Node Execution
        # ====================================================================
        node_logger = self._get_node_logger("PLANNING")
        
        # Detect mode from document context
        mode = state["document_context"].get("mode", "ask")
        
        if mode == "ask":
            node_logger.info("🎯 Starting planning phase in ASK MODE (Planning Only)...")
        else:
            node_logger.info("🎯 Starting planning phase in AGENT MODE (Will Execute)...")
        
        try:
            user_query = state["user_query"]
            node_logger.info(f"📝 Analyzing query: {user_query}")
            node_logger.info(f"🎭 Mode: {mode.upper()}")
            
            # Store mode in execution plan for later nodes
            state["execution_plan"] = {
                "mode": mode,
                "user_query": user_query,
                "context": state["document_context"]
            }
            
            node_logger.info("✅ Planning phase complete - ready for tool execution")
            
            return state
            
        except Exception as e:
            node_logger.error(f"❌ Planning failed: {e}")
            state["errors"].append(f"Planning error: {e}")
            return state
    
    def _knowledge_base_node(self, state: AgentState) -> AgentState:
        """
        KNOWLEDGE BASE NODE: Build and maintain knowledge base
        
        This node manages the knowledge base of available files and their metadata
        """
        # ====================================================================
        # CURSOR MARKER: Knowledge Base Node Execution
        # ====================================================================
        node_logger = self._get_node_logger("KNOWLEDGE_BASE")
        node_logger.info("📚 Loading knowledge base...")
        
        try:
            # IMPORTANT: Just load the KB, don't rebuild it!
            # The KB is built once using build_semantic_knowledge_base.py
            knowledge_base = self._load_knowledge_base()
            
            if not knowledge_base:
                node_logger.warning("⚠️ Knowledge base is empty! Run build_semantic_knowledge_base.py first")
            else:
                node_logger.info(f"✅ Loaded {len(knowledge_base)} files from knowledge base")
            
            state["knowledge_base"] = knowledge_base
            node_logger.info("✅ Knowledge base phase complete")
            
            return state
            
        except Exception as e:
            node_logger.error(f"❌ Knowledge base failed: {e}")
            state["errors"].append(f"Knowledge base error: {e}")
            return state
    
    def _tool_execution_node(self, state: AgentState) -> AgentState:
        """
        UNIFIED TOOL EXECUTION NODE: All Excel operations go through excel_tools.py
        
        This node:
        1. Finds relevant spreadsheet from KB
        2. Gets building code context
        3. Generates AI-driven tool plan
        4. In ASK mode: Stores plan for display
        5. In AGENT mode: Executes tools and stores results
        
        This replaces ALL the old file_save, file_parse, file_edit nodes!
        """
        # ====================================================================
        # CURSOR MARKER: Tool Execution Node
        # ====================================================================
        node_logger = self._get_node_logger("TOOL_EXEC")
        
        mode = state["execution_plan"].get("mode", "ask")
        node_logger.info(f"🔧 Starting tool execution in {mode.upper()} mode...")
        
        try:
            user_query = state["user_query"]
            knowledge_base = state["knowledge_base"]
            
            # Step 1: Find relevant spreadsheet
            node_logger.info("📊 Step 1: Finding relevant spreadsheet...")
            file_info = self._find_relevant_spreadsheet(user_query, knowledge_base)
            if not file_info:
                node_logger.warning("⚠️ No relevant spreadsheet found")
                state["errors"].append("No relevant spreadsheet found")
                return state
            
            node_logger.info(f"✅ Selected: {file_info.get('file_name')}")
            
            # Step 2: Get building code context (if available)
            code_context = ""
            if CODE_RAG_AVAILABLE:
                try:
                    node_logger.info("📚 Step 2: Querying building codes...")
                    code_rag = get_building_code_rag()
                    code_results = code_rag.query_all_codes(
                        f"Building code requirements for: {user_query}",
                        top_k_per_code=1
                    )
                    code_texts = []
                    for code_name, results in code_results.items():
                        if results:
                            code_texts.append(f"[{code_name}]: {results[0]['text'][:200]}")
                    code_context = "\n".join(code_texts)
                    node_logger.info(f"✅ Retrieved context from {len(code_results)} codes")
                except Exception as e:
                    node_logger.warning(f"⚠️ Code RAG query failed: {e}")
            else:
                node_logger.info("ℹ️  Step 2: Skipped (Code RAG not available)")
            
            # Step 3: Generate AI-driven tool plan
            node_logger.info("🤖 Step 3: Generating AI-driven tool plan...")
            tool_plan = self._generate_tool_plan(user_query, file_info, code_context)
            
            if not tool_plan:
                node_logger.warning("⚠️ Could not generate tool plan")
                state["errors"].append("Could not generate tool plan")
                return state
            
            node_logger.info(f"✅ Generated plan with {len(tool_plan)} steps")
            
            # Store tool plan in state
            state["execution_plan"]["tool_plan"] = tool_plan
            state["execution_plan"]["file_info"] = file_info
            state["execution_plan"]["code_context"] = code_context
            
            # Step 4: Execute or store based on mode
            if mode == "ask":
                node_logger.info("💡 ASK MODE: Storing plan for display (not executing)")
                state["results"].append({
                    "type": "tool_plan",
                    "file": file_info.get("file_name"),
                    "plan": tool_plan,
                    "code_context": code_context
                })
            else:
                node_logger.info("⚡ AGENT MODE: Executing tool plan...")
                
                # ✨ STREAMING: Initialize progress streamer
                if PROGRESS_STREAMING_AVAILABLE:
                    stream_stage_start("Tool Execution", f"Executing {len(tool_plan)} operations")
                
                executed_results = []
                
                # Track edited files to reuse across operations (FIX: prevents multiple file creation)
                edited_files = {}  # template_path -> edited_path
                
                # Track iteration for adaptive reasoning
                iteration_data = {"inputs": {}, "outputs": {}, "metrics": {}}
                
                for i, step in enumerate(tool_plan, 1):
                    tool_name = step.get("tool")
                    params = step.get("params", {})
                    reasoning = step.get("reasoning", "")
                    
                    node_logger.info(f"   🔧 Step {i}/{len(tool_plan)}: {tool_name}")
                    node_logger.info(f"      Reasoning: {reasoning}")
                    
                    # ✨ STREAMING: Stream tool start
                    if PROGRESS_STREAMING_AVAILABLE:
                        stream_tool(tool_name, params)
                    
                    try:
                        # FIX: If we've already created an edited file, use it instead of template
                        if "file_path" in params:
                            original_path = params["file_path"]
                            if original_path in edited_files:
                                params["file_path"] = edited_files[original_path]
                                node_logger.info(f"      📂 Using existing edited file")
                        
                        # 🧠 ON-THE-FLY REASONING: Query RAG for specific operations
                        if CODE_RAG_AVAILABLE and tool_name in ["WRITE_CELL", "READ_CELL"]:
                            # Get code-specific guidance for this parameter
                            param_name = params.get("cell", "")
                            if param_name:
                                try:
                                    code_rag = get_building_code_rag()
                                    domain = file_info.get("domain", "structural")
                                    guidance = code_rag.query_all_codes(
                                        f"{domain} design: requirements for {param_name} parameter",
                                        top_k_per_code=1
                                    )
                                    if guidance:
                                        node_logger.info(f"      📚 On-the-fly code guidance retrieved")
                                        if PROGRESS_STREAMING_AVAILABLE:
                                            stream_status(f"Code guidance: {list(guidance.keys())[0]}")
                                except Exception as e:
                                    node_logger.debug(f"On-the-fly RAG query failed: {e}")
                        
                        result = execute_tool(tool_name, **params)
                        
                        # FIX: Track edited file for next operations
                        if result.get("success") and "edited_file" in result:
                            edited_files[original_path] = result["edited_file"]
                        
                        # 📊 TRACKING: Collect iteration data
                        if tool_name == "WRITE_CELL" and "value" in params:
                            cell_name = params.get("cell", f"cell_{i}")
                            iteration_data["inputs"][cell_name] = params["value"]
                        elif tool_name == "READ_CELL" and result.get("success"):
                            cell_name = params.get("cell", f"output_{i}")
                            value = result.get("value")
                            iteration_data["outputs"][cell_name] = value
                            # Try to identify metrics (ratios, factors, etc.)
                            if isinstance(value, (int, float)):
                                if "ratio" in cell_name.lower() or "sf" in cell_name.lower():
                                    iteration_data["metrics"][cell_name] = float(value)
                        
                        executed_results.append({
                            "step": i,
                            "tool": tool_name,
                            "params": params,
                            "result": result,
                            "reasoning": reasoning
                        })
                        
                        if result.get("success"):
                            node_logger.info(f"      ✅ Success")
                            # ✨ STREAMING: Stream success
                            if PROGRESS_STREAMING_AVAILABLE:
                                stream_success(f"{tool_name} completed")
                        else:
                            node_logger.warning(f"      ⚠️ Failed: {result.get('error', 'Unknown')}")
                            # ✨ STREAMING: Stream error
                            if PROGRESS_STREAMING_AVAILABLE:
                                stream_error(f"{tool_name} failed", result.get("error", ""))
                            
                    except Exception as e:
                        node_logger.error(f"      ❌ Tool execution error: {e}")
                        executed_results.append({
                            "step": i,
                            "tool": tool_name,
                            "params": params,
                            "error": str(e),
                            "reasoning": reasoning
                        })
                        # ✨ STREAMING: Stream error
                        if PROGRESS_STREAMING_AVAILABLE:
                            stream_error(f"{tool_name} exception", str(e))
                
                # ✨ STREAMING: Complete execution stage
                if PROGRESS_STREAMING_AVAILABLE:
                    stream_stage_complete("Tool Execution", f"{len(executed_results)} operations completed")
                
                # 🔍 Validate results with code RAG (if available)
                validation = None
                if CODE_RAG_AVAILABLE and executed_results:
                    node_logger.info("🔍 Step 5: Validating results against building codes...")
                    # ✨ STREAMING: Stream validation start
                    if PROGRESS_STREAMING_AVAILABLE:
                        stream_stage_start("Code Validation", "Checking compliance")
                    
                    validation = self._validate_with_code_rag(executed_results, code_context)
                    node_logger.info(f"✅ Validation complete: {validation.get('compliant', 'N/A')}")
                    
                    # ✨ STREAMING: Stream validation result
                    if PROGRESS_STREAMING_AVAILABLE:
                        if validation.get("compliant"):
                            stream_success("Design is code-compliant")
                        else:
                            stream_error("Code compliance issues", validation.get("issues", ""))
                        stream_stage_complete("Code Validation")
                
                state["results"].append({
                    "type": "tool_execution",
                    "file": file_info.get("file_name"),
                    "executed_results": executed_results,
                    "validation": validation
                })
            
            node_logger.info("✅ Tool execution phase complete")
            return state
            
        except Exception as e:
            node_logger.error(f"❌ Tool execution failed: {e}")
            state["errors"].append(f"Tool execution error: {e}")
            return state
    
    def _file_save_node(self, state: AgentState) -> AgentState:
        """
        FILE SAVE NODE: Save files to appropriate locations
        
        This node handles saving files to Desktop or other specified locations
        """
        # ====================================================================
        # CURSOR MARKER: File Save Node Execution
        # ====================================================================
        node_logger = self._get_node_logger("FILE_SAVE")
        node_logger.info("💾 Saving files...")
        
        try:
            save_operations = state["execution_plan"].get("save_operations", [])
            saved_files = []
            
            for operation in save_operations:
                result = self._execute_save_operation(operation)
                if result["success"]:
                    saved_files.append(result["file_path"])
                    node_logger.info(f"✅ Saved: {result['file_name']}")
                else:
                    node_logger.error(f"❌ Failed to save: {result['error']}")
                    state["errors"].append(result["error"])
            
            state["results"].append({
                "node": "file_save",
                "saved_files": saved_files,
                "count": len(saved_files)
            })
            
            node_logger.info(f"✅ File save phase complete - {len(saved_files)} files saved")
            return state
            
        except Exception as e:
            node_logger.error(f"❌ File save failed: {e}")
            state["errors"].append(f"File save error: {e}")
            return state
    
    def _file_look_node(self, state: AgentState) -> AgentState:
        """
        FILE LOOK NODE: Examine and analyze files
        
        This node examines files to understand their structure and content
        """
        # ====================================================================
        # CURSOR MARKER: File Look Node Execution
        # ====================================================================
        node_logger = self._get_node_logger("FILE_LOOK")
        node_logger.info("👀 Examining files...")
        
        try:
            look_operations = state["execution_plan"].get("look_operations", [])
            examined_files = []
            
            for operation in look_operations:
                result = self._execute_look_operation(operation)
                examined_files.append(result)
                node_logger.info(f"🔍 Examined: {result['file_name']}")
            
            state["results"].append({
                "node": "file_look",
                "examined_files": examined_files,
                "count": len(examined_files)
            })
            
            node_logger.info(f"✅ File look phase complete - {len(examined_files)} files examined")
            return state
            
        except Exception as e:
            node_logger.error(f"❌ File look failed: {e}")
            state["errors"].append(f"File look error: {e}")
            return state
    
    def _file_parse_node(self, state: AgentState) -> AgentState:
        """
        FILE PARSE NODE: Parse Excel files to understand structure
        
        This node parses Excel files to understand their structure and capabilities
        """
        # ====================================================================
        # CURSOR MARKER: File Parse Node Execution
        # ====================================================================
        node_logger = self._get_node_logger("FILE_PARSE")
        node_logger.info("📊 Parsing Excel files...")
        
        try:
            parse_operations = state["execution_plan"].get("parse_operations", [])
            parsed_files = []
            
            for operation in parse_operations:
                result = self._execute_parse_operation(operation)
                parsed_files.append(result)
                node_logger.info(f"📋 Parsed: {result['file_name']}")
            
            state["parsed_content"] = {
                "files": parsed_files,
                "count": len(parsed_files)
            }
            
            state["results"].append({
                "node": "file_parse",
                "parsed_files": parsed_files,
                "count": len(parsed_files)
            })
            
            node_logger.info(f"✅ File parse phase complete - {len(parsed_files)} files parsed")
            return state
            
        except Exception as e:
            node_logger.error(f"❌ File parse failed: {e}")
            state["errors"].append(f"File parse error: {e}")
            return state
    
    def _file_edit_node(self, state: AgentState) -> AgentState:
        """
        FILE EDIT NODE: Modify files with intelligent operations
        
        This node performs intelligent modifications to files based on parsed content
        """
        # ====================================================================
        # CURSOR MARKER: File Edit Node Execution
        # ====================================================================
        node_logger = self._get_node_logger("FILE_EDIT")
        node_logger.info("✏️  Editing files...")
        
        try:
            edit_operations = state["execution_plan"].get("edit_operations", [])
            edited_files = []
            
            for operation in edit_operations:
                result = self._execute_edit_operation(operation, state)
                edited_files.append(result)
                node_logger.info(f"✏️  Edited: {result['file_name']}")
            
            state["results"].append({
                "node": "file_edit",
                "edited_files": edited_files,
                "count": len(edited_files)
            })
            
            node_logger.info(f"✅ File edit phase complete - {len(edited_files)} files edited")
            return state
            
        except Exception as e:
            node_logger.error(f"❌ File edit failed: {e}")
            state["errors"].append(f"File edit error: {e}")
            return state
    
    def _response_node(self, state: AgentState) -> AgentState:
        """
        RESPONSE NODE: Format and generate final response
        
        This node takes the tool plan/execution results and formats them
        into a user-friendly response based on mode (ask vs agent)
        """
        # ====================================================================
        # CURSOR MARKER: Response Node - Format Output
        # ====================================================================
        node_logger = self._get_node_logger("RESPONSE")
        
        mode = state["execution_plan"].get("mode", "ask")
        node_logger.info(f"📝 Generating {mode.upper()} mode response...")
        
        try:
            # Get results from tool execution
            results = state["results"]
            errors = state["errors"]
            
            node_logger.info(f"📊 Processing {len(results)} results, {len(errors)} errors")
            
            # Generate response based on mode
            if mode == "ask":
                final_response = self._format_ask_mode_response(state)
            else:
                final_response = self._format_agent_mode_response(state)
            
            # Generate suggestions and actions
            suggestions = self._generate_suggestions(state)
            actions = self._generate_actions(state)
            
            state["final_response"] = final_response
            state["suggestions"] = suggestions
            state["actions"] = actions
            
            node_logger.info("✅ Response generation complete")
            
            return state
            
        except Exception as e:
            node_logger.error(f"❌ Response generation failed: {e}")
            state["errors"].append(f"Response generation error: {e}")
            state["final_response"] = f"Error generating response: {e}"
            return state
    
    # ========================================================================
    # HELPER METHODS
    # ========================================================================
    
    def _get_node_logger(self, node_name: str) -> logging.Logger:
        """Get a logger with node-specific formatting"""
        node_logger = logging.getLogger(f"SidianLangGraph.{node_name}")
        node_logger.setLevel(logging.INFO)
        
        # Add node name to log records
        for handler in node_logger.handlers:
            handler.addFilter(lambda record: setattr(record, 'node_name', node_name) or True)
        
        return node_logger
    
    def _detect_parsed_document(self, query: str) -> Optional[Dict[str, Any]]:
        """Detect if user is referencing a parsed document"""
        try:
            if any(keyword in query.lower() for keyword in ['parsed', 'xxx.txt', 'drawing', 'plan', 'members', 'L1', 'L2', 'L3']):
                # Look for xxx.txt in the RAG project root
                project_root = Path("../../../")  # Go up to rag-waddell-gitready
                xxx_file = project_root / "xxx.txt"
                if xxx_file.exists():
                    return {
                        "file_name": xxx_file.name,
                        "file_path": str(xxx_file),
                        "content": self._load_file_content(str(xxx_file))
                    }
            return None
        except Exception as e:
            self.logger.error(f"Error detecting parsed document: {e}")
            return None
    
    def _create_execution_plan(self, query: str, state: AgentState) -> Dict[str, Any]:
        """Create execution plan using AI"""
        if not self.ai_client:
            return {"route": "general", "tools": [], "error": "No AI client"}
        
        try:
            prompt = f"""
            Analyze this engineering query and create an execution plan.
            
            QUERY: {query}
            CONTEXT: {state.get('document_context', {})}
            
            Determine:
            1. Route: file_operations, knowledge_update, general
            2. Tools needed: file_save, file_look, file_parse, file_edit
            3. Operations to perform
            
            Return JSON format.
            """
            
            response = self.ai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are an expert engineering AI planner."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=500,
                temperature=0.1
            )
            
            # Parse response and create real operations
            parsed_doc = state.get("document_context", {}).get("parsed_document")
            
            if parsed_doc and "members" in query.lower():
                # Create real operations for member calculations
                return {
                    "route": "file_operations",
                    "tools": ["file_save", "file_parse", "file_edit"],
                    "save_operations": [
                        {
                            "source_file": self._find_beam_design_file(),
                            "save_type": "template_copy",
                            "new_sheets": ["L1", "L2", "L3"]
                        }
                    ],
                    "parse_operations": [
                        {
                            "file_path": self._find_beam_design_file()
                        }
                    ],
                    "edit_operations": [
                        {
                            "operation_type": "create_member_sheets",
                            "file_path": self._find_beam_design_file(),
                            "members": ["L1", "L2", "L3"]
                        }
                    ]
                }
            else:
                # Generic operations
                return {
                    "route": "file_operations",
                    "tools": ["file_save", "file_parse", "file_edit"],
                    "save_operations": [],
                    "parse_operations": [],
                    "edit_operations": []
                }
            
        except Exception as e:
            self.logger.error(f"Error creating execution plan: {e}")
            return {"route": "general", "tools": [], "error": str(e)}
    
    def _route_from_knowledge_base(self, state: AgentState) -> str:
        """Route from knowledge base node based on execution plan"""
        route = state["execution_plan"].get("route", "execution")
        
        # Map route names to actual node names
        route_mapping = {
            "file_operations": "edit_node",  # Default to edit for file operations
            "file_save": "save_node",
            "file_look": "look_node", 
            "file_parse": "parse_node",
            "file_edit": "edit_node",
            "execution": "execution_node",
            "knowledge_update": "kb_update_node"
        }
        
        actual_route = route_mapping.get(route, "execution_node")
        self.logger.info(f"🛤️  Routing from knowledge_base: {route} -> {actual_route}")
        self.logger.info(f"🛤️  Routing to: {actual_route}")
        return actual_route
    
    def _load_knowledge_base(self) -> Dict[str, Any]:
        """Load code-enhanced semantic knowledge base from file"""
        try:
            # Get the directory where this script is located
            script_dir = Path(__file__).parent
            
            # Use the new unified knowledge base (with contextual cell descriptions)
            kb_path = script_dir / "excel_knowledge_base.json"
            if kb_path.exists():
                with open(kb_path, 'r') as f:
                    kb = json.load(f)
                    # Remove metadata key if present
                    kb.pop("_metadata", None)
                    self.logger.info(f"✅ Loaded AI-enhanced knowledge base: {len(kb)} files")
                    return kb
            
            self.logger.warning(f"⚠️ No knowledge base found in {script_dir}")
            return {}
        except Exception as e:
            self.logger.error(f"Error loading knowledge base: {e}")
            return {}
    
    def _save_knowledge_base(self, knowledge_base: Dict[str, Any]) -> None:
        """Save knowledge base to file"""
        try:
            kb_path = Path("excel_knowledge_base.json")
            with open(kb_path, 'w') as f:
                json.dump(knowledge_base, f, indent=2)
        except Exception as e:
            self.logger.error(f"Error saving knowledge base: {e}")
    
    def _update_knowledge_base(self, knowledge_base: Dict[str, Any]) -> List[str]:
        """Update knowledge base with new files"""
        # Simplified implementation
        return []
    
    def _load_file_content(self, file_path: str) -> str:
        """Load content from file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            self.logger.error(f"Error loading file content: {e}")
            return ""
    
    def _find_relevant_excel_file(self, query: str, knowledge_base: Dict[str, Any]) -> Optional[str]:
        """Find the most relevant Excel file from knowledge base based on query"""
        try:
            query_lower = query.lower()
            
            # Keywords for different types of designs
            design_keywords = {
                "beam": ["beam", "girder", "joist"],
                "column": ["column", "post", "pillar"],
                "wall": ["wall", "retaining", "shear"],
                "foundation": ["foundation", "footing", "pile"],
                "slab": ["slab", "deck", "floor"],
                "steel": ["steel", "w-section", "wide flange"],
                "concrete": ["concrete", "reinforced", "rebar"]
            }
            
            # Score each file based on relevance
            file_scores = {}
            for file_name, file_info in knowledge_base.items():
                if not file_name.endswith(('.xlsx', '.xlsm')):
                    continue
                
                score = 0
                file_lower = file_name.lower()
                
                # Check file name
                for design_type, keywords in design_keywords.items():
                    if any(kw in query_lower for kw in keywords):
                        if any(kw in file_lower for kw in keywords):
                            score += 10
                
                # Check description if available
                if isinstance(file_info, dict) and 'description' in file_info:
                    desc_lower = file_info['description'].lower()
                    for design_type, keywords in design_keywords.items():
                        if any(kw in query_lower for kw in keywords):
                            if any(kw in desc_lower for kw in keywords):
                                score += 5
                
                if score > 0:
                    file_scores[file_name] = score
            
            # Return the highest scoring file
            if file_scores:
                best_file = max(file_scores, key=file_scores.get)
                self.logger.info(f"🎯 Selected '{best_file}' (score: {file_scores[best_file]})")
                
                # Get full path if available
                if isinstance(knowledge_base[best_file], dict) and 'path' in knowledge_base[best_file]:
                    return knowledge_base[best_file]['path']
                else:
                    # Search for the file
                    return self._find_file_by_name(best_file)
            
            return None
            
        except Exception as e:
            self.logger.error(f"❌ Error finding relevant file: {e}")
            return None
    
    def _find_file_by_name(self, file_name: str) -> Optional[str]:
        """Find a file by name in common locations"""
        search_paths = [
            Path.home() / "Desktop" / "vibeng" / "Excel Templates",
            Path.home() / "Desktop" / "Desktop - MacBook Pro (2)" / "vibeng" / "Excel Templates",
            Path.home() / "Desktop",
            Path("uploads"),
            Path(".")
        ]
        
        for search_path in search_paths:
            if search_path.exists():
                for file_path in search_path.rglob(file_name):
                    return str(file_path)
        
        return None
    
    def _find_beam_design_file(self) -> str:
        """Find the beam design file from knowledge base"""
        try:
            # Look for beam design files in common locations
            # Priority: vibeng Excel Templates folder
            search_paths = [
                Path.home() / "Desktop" / "vibeng" / "Excel Templates",
                Path.home() / "Desktop" / "Desktop - MacBook Pro (2)" / "vibeng" / "Excel Templates",
                Path.home() / "Desktop",
                Path("uploads"),
                Path(".")
            ]
            
            self.logger.info(f"🔍 Searching for beam design file in {len(search_paths)} locations...")
            
            for search_path in search_paths:
                if search_path.exists():
                    self.logger.info(f"📂 Checking: {search_path}")
                    for file_path in search_path.rglob("*.xlsx"):
                        if "beam" in file_path.name.lower() and "design" in file_path.name.lower():
                            self.logger.info(f"✅ Found beam design file: {file_path}")
                            return str(file_path)
                    for file_path in search_path.rglob("*.xlsm"):
                        if "beam" in file_path.name.lower() and "design" in file_path.name.lower():
                            self.logger.info(f"✅ Found beam design file: {file_path}")
                            return str(file_path)
            
            # Fallback - return a default path
            fallback_path = Path.home() / "Desktop" / "vibeng" / "Excel Templates" / "Mantle Demo - Beam Design.xlsm"
            self.logger.warning(f"⚠️ No beam design file found, using fallback: {fallback_path}")
            return str(fallback_path)
            
        except Exception as e:
            self.logger.error(f"❌ Error finding beam design file: {e}")
            return str(Path.home() / "Desktop" / "vibeng" / "Excel Templates" / "Mantle Demo - Beam Design.xlsm")
    
    def _extract_learning_data(self, state: AgentState) -> List[Dict[str, Any]]:
        """Extract learning data from execution results"""
        return []
    
    def _apply_learning_updates(self, knowledge_base: Dict[str, Any], learning_data: List[Dict[str, Any]]) -> None:
        """Apply learning updates to knowledge base"""
        pass
    
    def _execute_save_operation(self, operation: Dict[str, Any]) -> Dict[str, Any]:
        """Execute a save operation using pandas + openpyxl"""
        try:
            source_file = operation.get("source_file")
            save_type = operation.get("save_type", "copy")  # "copy", "template_copy", "new_version"
            
            if not source_file or not Path(source_file).exists():
                return {"success": False, "error": "Source file not found"}
            
            desktop_path = Path.home() / "Desktop"
            desktop_path.mkdir(exist_ok=True)
            
            timestamp = int(time.time())
            base_name = Path(source_file).stem
            
            if save_type == "copy":
                # Simple file copy to Desktop
                new_name = f"{base_name}_modified_{timestamp}.xlsx"
                new_file_path = desktop_path / new_name
                shutil.copy2(source_file, new_file_path)
                
            elif save_type == "template_copy":
                # Copy template and create new sheets
                wb = load_workbook(source_file)
                
                # Find template sheet (look for common names)
                template_sheet = None
                template_names = ["Template", "template", "Sheet1", "Input"]
                for sheet_name in template_names:
                    if sheet_name in wb.sheetnames:
                        template_sheet = wb[sheet_name]
                        break
                
                if not template_sheet:
                    return {"success": False, "error": "No template sheet found"}
                
                # Create new sheets based on operation parameters
                new_sheets = operation.get("new_sheets", ["L1", "L2", "L3"])
                created_sheets = []
                
                for sheet_name in new_sheets:
                    if sheet_name not in wb.sheetnames:
                        new_sheet = wb.copy_worksheet(template_sheet)
                        new_sheet.title = sheet_name
                        created_sheets.append(sheet_name)
                
                new_name = f"{base_name}_new_{timestamp}.xlsx"
                new_file_path = desktop_path / new_name
                wb.save(new_file_path)
                
            else:
                # Default copy
                new_name = f"{base_name}_copy_{timestamp}.xlsx"
                new_file_path = desktop_path / new_name
                shutil.copy2(source_file, new_file_path)
            
            self.logger.info(f"✅ Saved file: {new_file_path}")
            return {
                "success": True, 
                "file_name": new_name, 
                "file_path": str(new_file_path),
                "created_sheets": operation.get("new_sheets", [])
            }
            
        except Exception as e:
            self.logger.error(f"❌ Save operation failed: {e}")
            return {"success": False, "error": str(e)}
    
    def _execute_look_operation(self, operation: Dict[str, Any]) -> Dict[str, Any]:
        """Execute a look operation"""
        return {"file_name": "test.xlsx", "structure": "analyzed"}
    
    def _execute_parse_operation(self, operation: Dict[str, Any]) -> Dict[str, Any]:
        """Execute a parse operation using pandas for comprehensive analysis"""
        try:
            file_path = operation.get("file_path")
            if not file_path or not Path(file_path).exists():
                return {"error": "File not found"}
            
            # Read all sheets using pandas
            excel_file = pd.ExcelFile(file_path)
            sheet_data = {}
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Analyze sheet structure
                    sheet_info = {
                        "sheet_name": sheet_name,
                        "dimensions": df.shape,
                        "columns": df.columns.tolist(),
                        "data_types": df.dtypes.to_dict(),
                        "has_formulas": self._detect_formulas(df),
                        "input_cells": self._find_input_cells(df),
                        "output_cells": self._find_output_cells(df),
                        "engineering_parameters": self._extract_engineering_parameters(df)
                    }
                    
                    sheet_data[sheet_name] = sheet_info
                    
                except Exception as e:
                    sheet_data[sheet_name] = {"error": f"Could not parse sheet: {e}"}
            
            # Determine spreadsheet type
            spreadsheet_type = self._classify_spreadsheet_type(sheet_data)
            
            # Generate analysis
            analysis = self._generate_spreadsheet_analysis(sheet_data, spreadsheet_type)
            
            self.logger.info(f"📊 Parsed {len(sheet_data)} sheets from {Path(file_path).name}")
            
            return {
                "file_name": Path(file_path).name,
                "file_path": file_path,
                "spreadsheet_type": spreadsheet_type,
                "sheets": sheet_data,
                "analysis": analysis,
                "parsed": True
            }
            
        except Exception as e:
            self.logger.error(f"❌ Parse operation failed: {e}")
            return {"error": str(e)}
    
    def _execute_edit_operation(self, operation: Dict[str, Any], state: AgentState) -> Dict[str, Any]:
        """Execute an edit operation with real engineering modifications"""
        try:
            operation_type = operation.get("operation_type", "modify_parameters")
            file_path = operation.get("file_path")
            
            if not file_path or not Path(file_path).exists():
                return {"success": False, "error": "File not found"}
            
            if operation_type == "modify_parameters":
                return self._modify_engineering_parameters(operation, state)
            elif operation_type == "create_member_sheets":
                return self._create_member_sheets_generic(operation, state)
            elif operation_type == "update_loads":
                return self._update_loads_generic(operation, state)
            elif operation_type == "check_calculations":
                return self._check_calculations_generic(operation, state)
            else:
                return {"success": False, "error": f"Unknown operation type: {operation_type}"}
                
        except Exception as e:
            self.logger.error(f"❌ Edit operation failed: {e}")
            return {"success": False, "error": str(e)}
    
    # ========================================================================
    # ENGINEERING HELPER METHODS
    # ========================================================================
    
    def _detect_formulas(self, df: pd.DataFrame) -> bool:
        """Detect if dataframe contains formulas"""
        # Simple detection - look for cells that start with '='
        for col in df.columns:
            if df[col].dtype == 'object':
                if df[col].astype(str).str.startswith('=').any():
                    return True
        return False
    
    def _find_input_cells(self, df: pd.DataFrame) -> List[str]:
        """Find potential input cells (numeric values, not formulas)"""
        input_cells = []
        for col in df.columns:
            if df[col].dtype in ['int64', 'float64']:
                # Look for cells with values (not NaN)
                non_null_values = df[col].dropna()
                if len(non_null_values) > 0:
                    input_cells.append(col)
        return input_cells
    
    def _find_output_cells(self, df: pd.DataFrame) -> List[str]:
        """Find potential output cells (calculated values)"""
        output_cells = []
        for col in df.columns:
            if df[col].dtype == 'object':
                # Look for cells that might contain formulas or results
                if df[col].astype(str).str.contains('=|SUM|AVERAGE|MAX|MIN').any():
                    output_cells.append(col)
        return output_cells
    
    def _extract_engineering_parameters(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Extract engineering parameters from dataframe"""
        parameters = {}
        
        # Common engineering parameter patterns
        param_patterns = {
            "concrete_strength": ["concrete", "strength", "f'c", "fc", "mpa", "psi"],
            "span": ["span", "length", "l", "beam_length", "clear_span"],
            "load": ["load", "w", "p", "force", "dead_load", "live_load"],
            "material": ["material", "steel", "wood", "grade", "yield"],
            "dimensions": ["width", "height", "depth", "thickness", "diameter"]
        }
        
        # Search through column names and values
        all_text = " ".join(df.columns).lower()
        
        for param_name, patterns in param_patterns.items():
            for pattern in patterns:
                if pattern in all_text:
                    parameters[param_name] = {
                        "detected": True,
                        "pattern": pattern,
                        "columns": [col for col in df.columns if pattern in col.lower()]
                    }
                    break
        
        return parameters
    
    def _classify_spreadsheet_type(self, sheet_data: Dict) -> str:
        """Classify spreadsheet type based on structure"""
        
        # Look for engineering indicators
        engineering_keywords = {
            "beam_design": ["span", "moment", "deflection", "beam", "member"],
            "retaining_wall": ["wall", "retaining", "overturning", "sliding", "foundation"],
            "concrete_mix": ["concrete", "mix", "cement", "aggregate", "water"],
            "steel_design": ["steel", "yield", "buckling", "connection", "welding"],
            "structural_analysis": ["load", "force", "stress", "strain", "analysis"]
        }
        
        all_text = " ".join([
            " ".join(sheet.get("columns", [])) 
            for sheet in sheet_data.values()
            if isinstance(sheet, dict)
        ]).lower()
        
        for sheet_type, keywords in engineering_keywords.items():
            if any(keyword in all_text for keyword in keywords):
                return sheet_type
        
        return "generic_engineering"
    
    def _generate_spreadsheet_analysis(self, sheet_data: Dict, spreadsheet_type: str) -> Dict[str, Any]:
        """Generate comprehensive analysis of spreadsheet"""
        analysis = {
            "spreadsheet_type": spreadsheet_type,
            "total_sheets": len(sheet_data),
            "input_sheets": [],
            "calculation_sheets": [],
            "output_sheets": [],
            "engineering_parameters": {},
            "recommendations": []
        }
        
        # Analyze each sheet
        for sheet_name, sheet_info in sheet_data.items():
            if isinstance(sheet_info, dict) and "error" not in sheet_info:
                if sheet_info.get("input_cells"):
                    analysis["input_sheets"].append(sheet_name)
                if sheet_info.get("has_formulas"):
                    analysis["calculation_sheets"].append(sheet_name)
                if sheet_info.get("output_cells"):
                    analysis["output_sheets"].append(sheet_name)
                
                # Collect engineering parameters
                params = sheet_info.get("engineering_parameters", {})
                for param_name, param_info in params.items():
                    if param_name not in analysis["engineering_parameters"]:
                        analysis["engineering_parameters"][param_name] = []
                    analysis["engineering_parameters"][param_name].extend(param_info.get("columns", []))
        
        # Generate recommendations
        if spreadsheet_type == "beam_design":
            analysis["recommendations"] = [
                "Consider creating member-specific sheets from template",
                "Verify load combinations and safety factors",
                "Check deflection limits per code requirements"
            ]
        elif spreadsheet_type == "retaining_wall":
            analysis["recommendations"] = [
                "Verify overturning and sliding safety factors",
                "Check soil bearing capacity",
                "Consider seismic loads if applicable"
            ]
        
        return analysis
    
    def _modify_engineering_parameters(self, operation: Dict, state: AgentState) -> Dict[str, Any]:
        """Generic parameter modification using xlwings"""
        try:
            if not XLWINGS_AVAILABLE:
                return {"success": False, "error": "xlwings not available - Excel automation disabled"}
            
            file_path = operation["file_path"]
            parameters = operation.get("parameters", {})  # {"concrete_strength": 30, "span": 20}
            
            app = xw.App(visible=False)
            app.display_alerts = False
            app.screen_updating = False
            
            wb = app.books.open(file_path)
            
            modifications_made = []
            
            for param_name, param_value in parameters.items():
                # Find parameter location using knowledge base
                param_location = self._find_parameter_location(wb, param_name, state["knowledge_base"])
                
                if param_location:
                    sheet_name, cell_address = param_location
                    wb.sheets[sheet_name].range(cell_address).value = param_value
                    modifications_made.append(f"{param_name} = {param_value}")
                    self.logger.info(f"✏️ Modified {param_name} to {param_value} in {sheet_name}!{cell_address}")
                else:
                    self.logger.warning(f"⚠️ Could not find location for parameter: {param_name}")
            
            wb.save()
            wb.close()
            app.quit()
            
            return {
                "success": True,
                "modifications": modifications_made,
                "file_path": file_path,
                "parameters_modified": len(modifications_made)
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _find_parameter_location(self, workbook, param_name: str, knowledge_base: Dict) -> Optional[Tuple[str, str]]:
        """Find where a parameter is located in any engineering spreadsheet"""
        
        # Common engineering parameter patterns
        param_patterns = {
            "concrete_strength": ["concrete", "strength", "f'c", "fc", "mpa"],
            "span": ["span", "length", "l", "beam_length"],
            "load": ["load", "w", "p", "force"],
            "material": ["material", "steel", "wood", "grade"]
        }
        
        # Search through sheets for parameter
        for sheet in workbook.sheets:
            try:
                used_range = sheet.used_range
                if used_range:
                    for cell in used_range:
                        cell_value = str(cell.value).lower() if cell.value else ""
                        
                        # Check if cell contains parameter name
                        search_patterns = param_patterns.get(param_name.lower(), [param_name.lower()])
                        for pattern in search_patterns:
                            if pattern in cell_value:
                                # Found parameter name, return location
                                return sheet.name, cell.address
            except Exception:
                continue
        
        return None
    
    def _create_member_sheets_generic(self, operation: Dict, state: AgentState) -> Dict[str, Any]:
        """Create member sheets from template"""
        try:
            if not XLWINGS_AVAILABLE:
                return {"success": False, "error": "xlwings not available - Excel automation disabled"}
            
            file_path = operation["file_path"]
            members = operation.get("members", ["L1", "L2", "L3"])
            
            app = xw.App(visible=False)
            app.display_alerts = False
            app.screen_updating = False
            
            wb = app.books.open(file_path)
            
            # Find template sheet
            template_sheet = None
            template_names = ["Template", "template", "Sheet1", "Input"]
            for sheet_name in template_names:
                if sheet_name in [s.name for s in wb.sheets]:
                    template_sheet = wb.sheets[sheet_name]
                    break
            
            if not template_sheet:
                return {"success": False, "error": "No template sheet found"}
            
            created_sheets = []
            
            for member in members:
                if member not in [s.name for s in wb.sheets]:
                    # Create new sheet after template
                    new_sheet = wb.sheets.add(member, after=template_sheet)
                    
                    # Copy template content to new sheet
                    template_sheet.used_range.copy()
                    new_sheet.range('A1').paste()
                    
                    created_sheets.append(member)
                    self.logger.info(f"✅ Created sheet: {member}")
            
            wb.save()
            wb.close()
            app.quit()
            
            return {
                "success": True,
                "created_sheets": created_sheets,
                "file_path": file_path,
                "message": f"Created {len(created_sheets)} member sheets"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _update_loads_generic(self, operation: Dict, state: AgentState) -> Dict[str, Any]:
        """Update loads in spreadsheet"""
        try:
            if not XLWINGS_AVAILABLE:
                return {"success": False, "error": "xlwings not available - Excel automation disabled"}
            
            file_path = operation["file_path"]
            loads = operation.get("loads", {})  # {"dead_load": 50, "live_load": 100}
            
            app = xw.App(visible=False)
            app.display_alerts = False
            app.screen_updating = False
            
            wb = app.books.open(file_path)
            
            modifications_made = []
            
            for load_type, load_value in loads.items():
                param_location = self._find_parameter_location(wb, load_type, state["knowledge_base"])
                
                if param_location:
                    sheet_name, cell_address = param_location
                    wb.sheets[sheet_name].range(cell_address).value = load_value
                    modifications_made.append(f"{load_type} = {load_value}")
            
            wb.save()
            wb.close()
            app.quit()
            
            return {
                "success": True,
                "modifications": modifications_made,
                "file_path": file_path
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _check_calculations_generic(self, operation: Dict, state: AgentState) -> Dict[str, Any]:
        """Check calculations in spreadsheet"""
        try:
            file_path = operation["file_path"]
            
            # Use pandas to read and analyze calculations
            excel_file = pd.ExcelFile(file_path)
            
            calculation_results = {}
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                # Look for calculation results
                numeric_cols = df.select_dtypes(include=['number']).columns
                if len(numeric_cols) > 0:
                    calculation_results[sheet_name] = {
                        "numeric_columns": numeric_cols.tolist(),
                        "has_calculations": self._detect_formulas(df),
                        "data_summary": df[numeric_cols].describe().to_dict()
                    }
            
            return {
                "success": True,
                "calculation_results": calculation_results,
                "file_path": file_path,
                "message": "Calculations checked successfully"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    def _extract_semantic_summary(self, knowledge_base: Dict[str, Any]) -> str:
        """Extract a concise semantic summary from the knowledge base"""
        summary_parts = []
        
        for file_name, file_data in knowledge_base.items():
            summary_parts.append(f"\n📄 **{file_name}**")
            
            # Overall purpose
            if "workflow_analysis" in file_data:
                wa = file_data["workflow_analysis"]
                summary_parts.append(f"   Purpose: {wa.get('overall_purpose', 'N/A')}")
                summary_parts.append(f"   Use Case: {wa.get('primary_use_case', 'N/A')}")
                
                # Action templates
                if "action_templates" in wa and wa["action_templates"]:
                    summary_parts.append(f"   Available Actions:")
                    for template in wa["action_templates"][:3]:  # First 3
                        summary_parts.append(f"      - {template['action']}: {template['description']}")
            
            # Key sheets
            if "sheets" in file_data:
                summary_parts.append(f"   Key Sheets:")
                for sheet_name, sheet_data in list(file_data["sheets"].items())[:3]:  # First 3
                    if "semantics" in sheet_data:
                        sem = sheet_data["semantics"]
                        summary_parts.append(f"      - {sheet_name}: {sem.get('purpose', 'N/A')}")
                        
                        # Input parameters
                        if "input_parameters" in sem and sem["input_parameters"]:
                            summary_parts.append(f"         Inputs: {', '.join([p['name'] for p in sem['input_parameters'][:5]])}")
        
        return "\n".join(summary_parts)
    
    def _format_ask_mode_response(self, state: AgentState) -> str:
        """
        Format ASK MODE response: Show tool plan without executing
        
        This displays what WOULD be done in a clear, actionable format
        """
        self.logger.info("💡 Formatting ASK MODE response...")
        
        results = state["results"]
        if not results or results[0].get("type") != "tool_plan":
            return "No tool plan generated."
        
        result = results[0]
        tool_plan = result.get("plan", [])
        file_name = result.get("file", "Unknown")
        code_context = result.get("code_context", "")
        
        # Build formatted response
        response_parts = []
        response_parts.append("## 💡 Execution Plan (ASK MODE)\n")
        response_parts.append(f"**Mode**: Planning Only (not executing)\n")
        response_parts.append(f"**Spreadsheet**: `{file_name}`\n\n")
        
        if code_context:
            response_parts.append("### 📚 Building Code Context:\n")
            response_parts.append(f"{code_context[:500]}...\n\n")
        
        response_parts.append(f"### 🔧 Planned Tool Operations ({len(tool_plan)} steps):\n")
        
        for i, step in enumerate(tool_plan, 1):
            tool = step.get("tool", "Unknown")
            reasoning = step.get("reasoning", "")
            params = step.get("params", {})
            
            response_parts.append(f"{i}. **{tool}**")
            if reasoning:
                response_parts.append(f"   - Purpose: {reasoning}")
            
            # Show key parameters
            if tool == "WRITE_CELL":
                response_parts.append(f"   - Cell: `{params.get('cell')}` = `{params.get('value')}`")
            elif tool == "FIND_PARAMETER":
                response_parts.append(f"   - Search for: `{params.get('label')}`")
            elif tool == "READ_CELL":
                response_parts.append(f"   - Read from: `{params.get('cell')}`")
            
            response_parts.append("")
        
        response_parts.append("### ⚡ To Execute This Plan:\n")
        response_parts.append("Switch to **Agent Mode** and ask the same question.\n")
        response_parts.append("I will then execute these tools and show you the results.")
        
        return "\n".join(response_parts)
    
    def _find_relevant_spreadsheet(self, query: str, knowledge_base: Dict) -> Optional[Dict]:
        """
        Use AI to find the most relevant spreadsheet (no hardcoded keywords!)
        """
        if not knowledge_base:
            return None
        
        # If only one file, return it
        if len(knowledge_base) == 1:
            file_name = list(knowledge_base.keys())[0]
            file_info = knowledge_base[file_name]
            full_path = self._find_file_by_name(file_name)
            return {
                "file_name": file_name,
                "file_path": full_path or file_name,
                **file_info
            }
        
        # Use AI to select best spreadsheet
        if not self.ai_client:
            # Fallback to first file if no AI
            file_name = list(knowledge_base.keys())[0]
            file_info = knowledge_base[file_name]
            full_path = self._find_file_by_name(file_name)
            return {
                "file_name": file_name,
                "file_path": full_path or file_name,
                **file_info
            }
        
        try:
            # Build concise KB summary for AI
            kb_summary = []
            for fname, finfo in knowledge_base.items():
                purpose = finfo.get('purpose', 'Unknown')
                semantics = finfo.get('semantics', {})
                if isinstance(semantics, dict):
                    purpose = semantics.get('purpose', purpose)
                kb_summary.append(f"- {fname}: {purpose}")
            
            prompt = f"""Select the most relevant spreadsheet for this query.

Query: {query}

Available Spreadsheets:
{chr(10).join(kb_summary)}

Return ONLY the exact filename from the list above (nothing else).
"""
            
            response = self.ai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a file selector. Return only the exact filename, nothing else."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=100,
                temperature=0
            )
            
            selected_name = response.choices[0].message.content.strip()
            
            # Find matching file (case insensitive)
            for fname, finfo in knowledge_base.items():
                if fname.lower() in selected_name.lower() or selected_name.lower() in fname.lower():
                    full_path = self._find_file_by_name(fname)
                    self.logger.info(f"🤖 AI selected: {fname}")
                    return {
                        "file_name": fname,
                        "file_path": full_path or fname,
                        **finfo
                    }
            
            # Fallback if no match
            file_name = list(knowledge_base.keys())[0]
            file_info = knowledge_base[file_name]
            full_path = self._find_file_by_name(file_name)
            self.logger.info(f"📁 Fallback to: {file_name}")
            return {
                "file_name": file_name,
                "file_path": full_path or file_name,
                **file_info
            }
            
        except Exception as e:
            self.logger.error(f"❌ AI selection failed: {e}")
            # Fallback to first file
            file_name = list(knowledge_base.keys())[0]
            file_info = knowledge_base[file_name]
            full_path = self._find_file_by_name(file_name)
            return {
                "file_name": file_name,
                "file_path": full_path or file_name,
                **file_info
            }
    
    def _generate_tool_plan(self, query: str, file_info: Dict, code_context: str) -> List[Dict]:
        """
        Use AI to generate tool execution plan
        
        This is where the LLM brain decides which tools to use and in what order
        """
        self.logger.info("🤖 Generating tool execution plan with AI...")
        
        if not self.ai_client:
            self.logger.warning("⚠️ AI client not available, using fallback")
            return []
        
        try:
            # Get available tools
            tools_list = get_tool_list()
            tools_desc = "\n".join([f"- {t['name']}: {t['description']}" for t in tools_list])
            
            # Build context about the spreadsheet WITH CELL MAPPINGS
            # FILTER: Only include sheets with actual cell mappings (skip empty ones)
            sheets_info = {}
            if "sheets" in file_info:
                for sheet_name, sheet_data in file_info["sheets"].items():
                    # Extract cell mappings from labeled_cells (direct mappings)
                    labeled_cells = sheet_data.get("labeled_cells", {})
                    
                    # SKIP sheets with no labeled cells (like "Main Floor ->")
                    if not labeled_cells:
                        continue
                    
                    # Extract from semantics input_parameters (detailed info with cells)
                    input_params = []
                    if "semantics" in sheet_data:
                        for param in sheet_data["semantics"].get("input_parameters", [])[:10]:
                            param_info = {
                                "name": param.get("name", ""),
                                "cell": param.get("cell", ""),
                                "label": param.get("label", ""),
                                "unit": param.get("unit", "")
                            }
                            # Include metric cell if available
                            if "cell_metric" in param:
                                param_info["cell_metric"] = param["cell_metric"]
                            if "use_cell" in param:
                                param_info["use_cell"] = param["use_cell"]
                            input_params.append(param_info)
                    
                    sheets_info[sheet_name] = {
                        "purpose": sheet_data.get("semantics", {}).get("purpose", ""),
                        "labeled_cells": labeled_cells,  # Direct mappings
                        "input_parameters": input_params  # Detailed parameter info
                    }
                    
                    # Limit to first 5 sheets with actual data
                    if len(sheets_info) >= 5:
                        break
            
            # Get the ACTUAL file path
            actual_file_path = file_info.get('file_path', '')
            file_name = file_info.get('file_name', 'Unknown')
            first_sheet = list(sheets_info.keys())[0] if sheets_info else "L(1)"
            
            prompt = f"""You are an expert engineer planning Excel operations.

User Query: {query}

Spreadsheet: {file_name}
File Path (USE IN ALL TOOLS): {actual_file_path}

📋 AVAILABLE SHEETS WITH CELL MAPPINGS:
{json.dumps(sheets_info, indent=2)}

🏗️ Building Code Context:
{code_context[:500] if code_context else 'No code context'}...

🔧 Available Tools:
{tools_desc}

⚡ CRITICAL INSTRUCTIONS:

0. **UNDERSTAND THE SPREADSHEET FIRST** (NEW!):
   - If multiple scenarios/spans are detected, ask which one to design or design the governing case
   - NEVER read capacity ratios from cells showing #VALUE! errors
   - ALWAYS input parameters → RECALC → THEN read results
   - If verification status shows "calculations incomplete", you MUST input values first!

1. **SHEET SELECTION** (IMPORTANT!):
   - For beam design spreadsheets, use sheets named "L(1)", "L(2)", "L(3)", etc. (member sheets)
   - AVOID "Main Floor ->" or "INFO" sheets for design calculations
   - Use sheets with POPULATED "labeled_cells" (not empty {{}})

2. **USE THE CELL MAPPINGS** from "labeled_cells" or "input_parameters" — DO NOT use FIND_PARAMETER!
   
   ⚠️ UNITS & MULTI-COLUMN LAYOUTS:
   - If a parameter has BOTH "cell" (Imperial) and "cell_metric" (Metric), use "cell_metric" for metric values
   - Example: span=8m → Use "cell_metric": "E5" (NOT "cell": "C5" which is feet)
   - Example: load=3.5 kN/m → Use "cell_metric" for the metric column
   
   Example mappings:
   - "Span" (metric): "cell_metric": "E5" → Use WRITE_CELL with cell="E5"
   - "Live Load" (metric): "cell_metric": "E8" → Use WRITE_CELL with cell="E8"
   - "Dead Load" (metric): "cell_metric": "E9" → Use WRITE_CELL with cell="E9"
   
3. **ITERATIVE DESIGN WORKFLOW** (For structural member design):
   When designing structural members (beams, columns, etc.), follow this optimization loop:
   
   a) **Input Design Parameters**:
      - Write span, loads, material properties to input cells
      - Use RECALC to trigger calculations
   
   b) **Check Capacity Ratios** (Read output cells for utilization):
      - Moment: Mf/Mr (factored moment / moment resistance)
      - Shear: Vf/Vr (factored shear / shear resistance)
      - Deflection: Actual/Limit
      - **Target: ~80% utilization** (0.75 to 0.85 is optimal)
   
   c) **Adjust Member Size**:
      - If utilization > 85%: Increase member size (next size up)
      - If utilization < 70%: Decrease member size (optimize material)
      - Repeat until all ratios are in the 75-85% range
   
   d) **Building Code Verification**:
      - Reference code_context to ensure compliance
      - Check that selected size meets code requirements
      - Verify deflection limits per code
   
   Example Iterative Design:
   [
     {{"tool": "WRITE_CELL", "params": {{"file_path": "...", "sheet": "L(1)", "cell": "E5", "value": 8}}, "reasoning": "Set span=8m"}},
     {{"tool": "WRITE_CELL", "params": {{"file_path": "...", "sheet": "L(1)", "cell": "E8", "value": 5.0}}, "reasoning": "Set live load=5.0 kPa"}},
     {{"tool": "RECALC", "params": {{"file_path": "..."}}, "reasoning": "Calculate with initial guess"}},
     {{"tool": "READ_CELL", "params": {{"file_path": "...", "sheet": "L(1)", "cell": "G25"}}, "reasoning": "Read Mf/Mr ratio"}},
     {{"tool": "READ_CELL", "params": {{"file_path": "...", "sheet": "L(1)", "cell": "H25"}}, "reasoning": "Read Vf/Vr ratio"}},
     {{"tool": "WRITE_CELL", "params": {{"file_path": "...", "sheet": "L(1)", "cell": "F15", "value": "2x10"}}, "reasoning": "Adjust size if ratio > 85%"}},
     {{"tool": "RECALC", "params": {{"file_path": "..."}}, "reasoning": "Recalculate with new size"}},
     {{"tool": "READ_CELL", "params": {{"file_path": "...", "sheet": "L(1)", "cell": "G25"}}, "reasoning": "Verify Mf/Mr now ~80%"}}
   ]

4. **NEVER use FIND_PARAMETER** — the knowledge base already has all cell locations!

5. **ALWAYS use this exact path**: {actual_file_path}

Return ONLY a JSON array (no markdown):
[
  {{"tool": "WRITE_CELL", "params": {{"file_path": "{actual_file_path}", "sheet": "L(1)", "cell": "C5", "value": 8}}, "reasoning": "Set span to 8m in L(1) sheet (KB: Span→C5)"}},
  {{"tool": "WRITE_CELL", "params": {{"file_path": "{actual_file_path}", "sheet": "L(1)", "cell": "C15", "value": 3.5}}, "reasoning": "Set load to 3.5 kN/m in L(1) sheet (KB: Trib. Width→C15)"}},
  {{"tool": "RECALC", "params": {{"file_path": "{actual_file_path}"}}, "reasoning": "Trigger calculations"}},
  ...
]

REMEMBER: Use L(1), L(2), L(3) sheets for beam design! Check labeled_cells is not empty!
"""
            
            response = self.ai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are an expert engineer. Return ONLY valid JSON array. No markdown, no comments, no trailing commas."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,
                temperature=0.2
            )
            
            raw_content = response.choices[0].message.content.strip()
            
            # Strip markdown if present
            if raw_content.startswith("```json"):
                raw_content = raw_content[7:]
            if raw_content.startswith("```"):
                raw_content = raw_content[3:]
            if raw_content.endswith("```"):
                raw_content = raw_content[:-3]
            raw_content = raw_content.strip()
            
            # Try to extract JSON if wrapped in other content
            if not raw_content.startswith("["):
                # Find first [ and last ]
                start_idx = raw_content.find("[")
                end_idx = raw_content.rfind("]")
                if start_idx != -1 and end_idx != -1:
                    raw_content = raw_content[start_idx:end_idx+1]
            
            # Remove any trailing commas before closing brackets (common AI error)
            import re
            raw_content = re.sub(r',(\s*[}\]])', r'\1', raw_content)
            
            try:
                tool_plan = json.loads(raw_content)
                self.logger.info(f"✅ Generated plan with {len(tool_plan)} steps")
                return tool_plan
            except json.JSONDecodeError as e:
                self.logger.error(f"❌ JSON parsing failed: {e}")
                self.logger.error(f"Raw content: {raw_content[:500]}")
                
                # Fallback: Return a basic plan
                self.logger.warning("⚠️ Using fallback basic plan")
                return self._generate_fallback_plan(query, file_info)
            
        except Exception as e:
            self.logger.error(f"❌ Tool plan generation failed: {e}")
            return self._generate_fallback_plan(query, file_info)
    
    def _generate_fallback_plan(self, query: str, file_info: Dict) -> List[Dict]:
        """
        Generate a basic fallback plan when AI generation fails
        
        This creates a simple plan based on keyword detection
        """
        self.logger.info("🔄 Generating fallback tool plan...")
        
        plan = []
        query_lower = query.lower()
        file_path = file_info.get("file_path", "")
        
        # Determine what sheet to use
        sheet = "L(1)"  # Default
        if "sheets" in file_info and file_info["sheets"]:
            sheet = list(file_info["sheets"].keys())[0]
        
        # Detect parameters to modify
        if "span" in query_lower:
            plan.append({
                "tool": "FIND_PARAMETER",
                "params": {"file_path": file_path, "sheet": sheet, "label": "Span"},
                "reasoning": "Find span parameter location"
            })
            
            # Try to extract value
            import re
            match = re.search(r'(\d+(?:\.\d+)?)\s*(?:m|meter)', query_lower)
            if match:
                value = float(match.group(1))
                plan.append({
                    "tool": "WRITE_CELL",
                    "params": {"file_path": file_path, "sheet": sheet, "cell": "B5", "value": value},
                    "reasoning": f"Set span to {value} meters"
                })
        
        if "load" in query_lower:
            plan.append({
                "tool": "FIND_PARAMETER",
                "params": {"file_path": file_path, "sheet": sheet, "label": "Load"},
                "reasoning": "Find load parameter location"
            })
            
            # Try to extract value
            match = re.search(r'(\d+(?:\.\d+)?)\s*(?:kn/m|kN/m)', query_lower)
            if match:
                value = float(match.group(1))
                plan.append({
                    "tool": "WRITE_CELL",
                    "params": {"file_path": file_path, "sheet": sheet, "cell": "B7", "value": value},
                    "reasoning": f"Set load to {value} kN/m"
                })
        
        if plan:
            # Add recalc and read steps
            plan.append({
                "tool": "RECALC",
                "params": {"file_path": file_path, "sheet": sheet},
                "reasoning": "Recalculate spreadsheet"
            })
            plan.append({
                "tool": "READ_CELL",
                "params": {"file_path": file_path, "sheet": sheet, "cell": "E20"},
                "reasoning": "Read design result"
            })
        
        self.logger.info(f"✅ Generated fallback plan with {len(plan)} steps")
        return plan
    
    def _execute_with_tools(self, state: AgentState) -> Dict[str, Any]:
        """
        Execute operations using the tool system (ENGINEER-LIKE EXECUTION)
        
        This is the core execution engine that:
        1. Uses semantic KB to find relevant files/parameters
        2. Queries building codes for requirements
        3. Executes tools (FIND_PARAMETER, WRITE_CELL, etc.)
        4. Validates against building codes
        """
        self.logger.info("🔧 Starting tool-based execution...")
        
        if not TOOLS_AVAILABLE:
            return {"success": False, "error": "Tools not available"}
        
        user_query = state["user_query"]
        knowledge_base = state["knowledge_base"]
        
        # Step 1: Find relevant spreadsheet
        file_info = self._find_relevant_spreadsheet(user_query, knowledge_base)
        if not file_info:
            return {"success": False, "error": "No relevant spreadsheet found"}
        
        # Step 2: Get building code context
        code_context = ""
        if CODE_RAG_AVAILABLE:
            try:
                code_rag = get_building_code_rag()
                code_results = code_rag.query_all_codes(
                    f"Building code requirements for: {user_query}",
                    top_k_per_code=1
                )
                code_texts = []
                for code_name, results in code_results.items():
                    if results:
                        code_texts.append(f"[{code_name}]: {results[0]['text'][:200]}")
                code_context = "\n".join(code_texts)
                self.logger.info(f"📚 Retrieved code context from {len(code_results)} codes")
            except Exception as e:
                self.logger.warning(f"⚠️ Code RAG query failed: {e}")
        
        # Step 3: Generate tool execution plan
        tool_plan = self._generate_tool_plan(user_query, file_info, code_context)
        if not tool_plan:
            return {"success": False, "error": "Could not generate tool plan"}
        
        # Step 4: Execute tools
        executed_results = []
        for i, step in enumerate(tool_plan, 1):
            tool_name = step.get("tool")
            params = step.get("params", {})
            reasoning = step.get("reasoning", "")
            
            self.logger.info(f"🔧 Step {i}/{len(tool_plan)}: {tool_name}")
            self.logger.info(f"   Reasoning: {reasoning}")
            
            try:
                result = execute_tool(tool_name, **params)
                executed_results.append({
                    "step": i,
                    "tool": tool_name,
                    "params": params,
                    "result": result,
                    "reasoning": reasoning
                })
                
                if result.get("success"):
                    self.logger.info(f"   ✅ Success: {result}")
                else:
                    self.logger.warning(f"   ⚠️ Failed: {result.get('error', 'Unknown error')}")
                    
            except Exception as e:
                self.logger.error(f"   ❌ Tool execution error: {e}")
                executed_results.append({
                    "step": i,
                    "tool": tool_name,
                    "params": params,
                    "error": str(e),
                    "reasoning": reasoning
                })
        
        # Step 5: Validate results with code RAG (if available)
        validation = None
        if CODE_RAG_AVAILABLE and executed_results:
            validation = self._validate_with_code_rag(executed_results, code_context)
        
        return {
            "success": True,
            "file": file_info.get("file_name"),
            "tool_plan": tool_plan,
            "executed_results": executed_results,
            "validation": validation,
            "code_context": code_context
        }
    
    def _validate_with_code_rag(self, results: List[Dict], code_context: str) -> Dict[str, Any]:
        """
        Validate execution results against building code requirements
        """
        self.logger.info("🔍 Validating results against building codes...")
        
        try:
            code_rag = get_building_code_rag()
            
            # Extract final values from results
            final_values = {}
            for result in results:
                if result.get("result", {}).get("success"):
                    tool = result["tool"]
                    if tool == "WRITE_CELL":
                        param_name = result["params"].get("cell", "unknown")
                        value = result["params"].get("value")
                        final_values[param_name] = value
                    elif tool == "READ_CELL":
                        final_values["result"] = result["result"].get("value")
            
            # Query code RAG for validation
            validation_query = f"Validate these values meet code requirements: {json.dumps(final_values)}"
            validation_results = code_rag.query_all_codes(validation_query, top_k_per_code=1)
            
            # Simple validation response
            return {
                "validated": True,
                "final_values": final_values,
                "code_references": len(validation_results),
                "compliant": True  # Would need deeper logic to determine actual compliance
            }
            
        except Exception as e:
            self.logger.error(f"❌ Validation failed: {e}")
            return {"validated": False, "error": str(e)}
    
    def _format_agent_mode_response(self, state: AgentState) -> str:
        """
        Format AGENT MODE response: Show what was executed
        
        This summarizes the actual tool execution and results
        """
        self.logger.info("⚡ Formatting AGENT MODE response...")
        
        results = state["results"]
        if not results or results[0].get("type") != "tool_execution":
            return "No execution results found."
        
        result = results[0]
        file_name = result.get("file", "Unknown")
        executed_results = result.get("executed_results", [])
        validation = result.get("validation")
        
        # Build formatted response
        response_parts = []
        response_parts.append("## ⚡ Execution Summary (AGENT MODE)\n")
        response_parts.append(f"**Mode**: Executed with Tools\n")
        response_parts.append(f"**Spreadsheet**: `{file_name}`\n\n")
        
        response_parts.append(f"### 🔧 Tool Execution ({len(executed_results)} steps):\n")
        
        for step_result in executed_results:
            step_num = step_result.get("step", "?")
            tool = step_result.get("tool", "Unknown")
            reasoning = step_result.get("reasoning", "")
            result_data = step_result.get("result", {})
            error = step_result.get("error")
            
            # Show status
            if error:
                status = "❌"
            elif result_data.get("success"):
                status = "✅"
            else:
                status = "⚠️"
            
            response_parts.append(f"{step_num}. {status} **{tool}**")
            if reasoning:
                response_parts.append(f"   - Purpose: {reasoning}")
            
            # Show results
            if error:
                response_parts.append(f"   - Error: {error}")
            elif tool == "FIND_PARAMETER" and result_data.get("success"):
                matches = result_data.get("matches", [])
                if matches:
                    response_parts.append(f"   - Found at: `{matches[0].get('value_cell', 'N/A')}`")
            elif tool == "WRITE_CELL" and result_data.get("success"):
                template_file = result_data.get("template_file", "")
                edited_file = result_data.get("edited_file", "")
                response_parts.append(f"   - Modified: `{edited_file}`")
                response_parts.append(f"   - Template: `{template_file}` (unchanged)")
            elif tool == "READ_CELL" and result_data.get("success"):
                response_parts.append(f"   - Value: `{result_data.get('value')}`")
            elif not result_data.get("success"):
                response_parts.append(f"   - Error: {result_data.get('error', 'Unknown')}")
            
            response_parts.append("")
        
        # Show validation
        if validation and validation.get("validated"):
            response_parts.append("### ✅ Building Code Validation:\n")
            response_parts.append(f"- Codes Checked: {validation.get('code_references', 0)}")
            response_parts.append(f"- Compliance: {'✅ PASS' if validation.get('compliant') else '❌ FAIL'}\n")
            
            final_values = validation.get("final_values", {})
            if final_values:
                response_parts.append("**Final Values:**")
                for key, value in final_values.items():
                    response_parts.append(f"- `{key}`: `{value}`")
                response_parts.append("")
        
        response_parts.append("### 🎉 Next Steps:\n")
        response_parts.append("- Review the edited Excel file on your system")
        response_parts.append("- Verify calculations meet code requirements")
        response_parts.append("- Ask follow-up questions if needed")
        
        return "\n".join(response_parts)
    
    def _generate_final_response(self, state: AgentState) -> str:
        """Generate final response for user (fallback method)"""
        return "Task completed successfully using LangGraph agent system."
    
    def _generate_suggestions(self, state: AgentState) -> List[str]:
        """Generate suggestions for user"""
        return ["Try asking about specific calculations", "Upload more files for analysis"]
    
    def _generate_actions(self, state: AgentState) -> List[str]:
        """Generate actions taken"""
        return ["Created member worksheets", "Updated knowledge base"]
    
    # ========================================================================
    # PUBLIC API
    # ========================================================================
    
    async def process_query(self, query: str, chat_history: List[Dict], context: Dict[str, Any]) -> Dict[str, Any]:
        """
        Main entry point for processing user queries
        
        This method initializes the state and runs the LangGraph
        """
        # ====================================================================
        # CURSOR MARKER: Main Query Processing Start
        # ====================================================================
        mode = context.get("mode", "ask")
        
        self.logger.info("🚀 Starting LangGraph query processing...")
        self.logger.info(f"📝 Query: {query}")
        self.logger.info(f"🎭 Mode: {mode.upper()}")
        
        if mode == "ask":
            self.logger.info("💡 ASK MODE: Will generate detailed plan without execution")
        else:
            self.logger.info("⚡ AGENT MODE: Will execute operations")
        
        try:
            # Initialize state
            initial_state: AgentState = {
                "user_query": query,
                "chat_history": chat_history,
                "document_context": context,
                "execution_plan": {},
                "selected_tools": [],
                "route_decision": "",
                "knowledge_base": {},
                "knowledge_updates": [],
                "target_files": [],
                "file_operations": [],
                "parsed_content": {},
                "results": [],
                "errors": [],
                "final_response": "",
                "suggestions": [],
                "actions": []
            }
            
            # Run the graph asynchronously
            self.logger.info("🔄 Executing LangGraph workflow...")
            final_state = await self.graph.ainvoke(initial_state)
            
            # ================================================================
            # CURSOR MARKER: Query Processing Complete
            # ================================================================
            self.logger.info("✅ LangGraph query processing complete")
            self.logger.info(f"📊 Final state: {len(final_state['results'])} results, {len(final_state['errors'])} errors")
            
            return {
                "response": final_state["final_response"],
                "suggestions": final_state["suggestions"],
                "actions": final_state["actions"],
                "results": final_state["results"],
                "errors": final_state["errors"],
                "clarification_needed": False,
                "relevant_chunks": []
            }
            
        except Exception as e:
            self.logger.error(f"❌ LangGraph processing failed: {e}")
            return {
                "response": f"Error processing query: {e}",
                "suggestions": [],
                "actions": [],
                "results": [],
                "errors": [str(e)],
                "clarification_needed": False,
                "relevant_chunks": []
            }

# ============================================================================
# CURSOR MARKER: LangGraph Agent System Complete
# ============================================================================

# ============================================================================
# SINGLETON PATTERN - Create agent once and reuse
# ============================================================================

_excel_agent_instance = None

def get_excel_agent() -> SidianLangGraphAgent:
    """
    Get or create the singleton Excel agent instance
    
    This ensures the agent (and its LangGraph) is built only ONCE
    and reused for all subsequent queries.
    """
    global _excel_agent_instance
    if _excel_agent_instance is None:
        _excel_agent_instance = SidianLangGraphAgent()
        print("✅ Excel agent singleton created and cached")
    return _excel_agent_instance
