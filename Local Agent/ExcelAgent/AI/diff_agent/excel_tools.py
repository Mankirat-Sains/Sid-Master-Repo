#!/usr/bin/env python3
"""
Excel Agent Tool System

Provides atomic operations that the LLM brain can orchestrate to perform
complex spreadsheet manipulations. Based on SheetMind's BNF grammar approach.

Tool Categories:
1. READ - Extract data
2. WRITE - Modify data
3. CALCULATE - Trigger computations
4. STRUCTURE - Manage sheets
5. ANALYSIS - Understand content
6. SAFETY - Backup and validation

Each tool is:
- Atomic (does one thing well)
- Composable (can be chained)
- Safe (validates inputs)
- Logged (tracks all operations)
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
import logging
import json
from datetime import datetime
import shutil

# Conditional xlwings import
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except Exception as e:
    XLWINGS_AVAILABLE = False
    xw = None


# ============================================================================
# LOGGING SETUP
# ============================================================================

logger = logging.getLogger("ExcelTools")
logger.setLevel(logging.INFO)

if not logger.handlers:
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter(
        '%(asctime)s - [EXCEL_TOOLS] - %(levelname)s - %(message)s'
    ))
    logger.addHandler(console_handler)


# ============================================================================
# HELPER FUNCTIONS
# ============================================================================

def _create_edited_copy(file_path: str) -> Optional[str]:
    """
    Create an EDITED copy of the template file (leaves original untouched)
    
    The edited file is saved in the same directory as the template with:
    - Name: {FILENAME}_edited_YYYYMMDD_HHMMSS.xlsx
    - Location: Same folder as template
    
    Returns: Path to edited copy or None if failed
    """
    try:
        source_path = Path(file_path)
        if not source_path.exists():
            return None
        
        # Create edited copy in SAME directory as template
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        edited_name = f"{source_path.stem}_edited_{timestamp}{source_path.suffix}"
        edited_path = source_path.parent / edited_name
        
        # Copy template to edited file
        shutil.copy2(source_path, edited_path)
        
        logger.info(f"   📄 Created edited copy: {edited_name}")
        return str(edited_path)
        
    except Exception as e:
        logger.error(f"❌ Failed to create edited copy: {e}")
        return None


# ============================================================================
# TOOL REGISTRY
# ============================================================================

TOOL_REGISTRY = {}

def register_tool(name: str, description: str, parameters: Dict[str, str]):
    """Decorator to register tools with their metadata"""
    def decorator(func):
        TOOL_REGISTRY[name] = {
            "function": func,
            "description": description,
            "parameters": parameters
        }
        return func
    return decorator


# ============================================================================
# READ TOOLS - Extract data from spreadsheets
# ============================================================================

@register_tool(
    name="READ_CELL",
    description="Read a single cell value from a specific sheet",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "cell": "Cell address (e.g., 'B5')"
    }
)
def read_cell(file_path: str, sheet: str, cell: str) -> Dict[str, Any]:
    """Read a single cell value"""
    try:
        logger.info(f"📖 READ_CELL: {file_path} | {sheet} | {cell}")
        
        wb = load_workbook(file_path, data_only=True)
        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet}' not found"}
        
        ws = wb[sheet]
        value = ws[cell].value
        wb.close()
        
        logger.info(f"   ✅ Value: {value}")
        return {"success": True, "value": value, "cell": cell, "sheet": sheet}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


@register_tool(
    name="READ_TABLE",
    description="Read a table/range from a sheet as a DataFrame",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "header_row": "Row number for headers (0-based)",
        "columns": "Optional: List of column names to read",
        "row_filter": "Optional: Filter condition (e.g., 'Column > 10')"
    }
)
def read_table(file_path: str, sheet: str, header_row: int = 0, 
               columns: Optional[List[str]] = None, 
               row_filter: Optional[str] = None) -> Dict[str, Any]:
    """Read a table/range as a DataFrame"""
    try:
        logger.info(f"📖 READ_TABLE: {file_path} | {sheet} | header_row={header_row}")
        
        df = pd.read_excel(file_path, sheet_name=sheet, header=header_row)
        
        # Filter columns if specified
        if columns:
            df = df[columns]
        
        # Apply row filter if specified (basic support)
        if row_filter:
            # This is simplified - in production, use safer query parsing
            df = df.query(row_filter)
        
        logger.info(f"   ✅ Read {len(df)} rows x {len(df.columns)} columns")
        
        return {
            "success": True,
            "data": df.to_dict(orient="records"),
            "columns": df.columns.tolist(),
            "row_count": len(df),
            "sheet": sheet
        }
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


@register_tool(
    name="READ_RANGE",
    description="Read a specific cell range (e.g., 'A1:D10')",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "range": "Cell range (e.g., 'A1:D10')"
    }
)
def read_range(file_path: str, sheet: str, range: str) -> Dict[str, Any]:
    """Read a cell range"""
    try:
        logger.info(f"📖 READ_RANGE: {file_path} | {sheet} | {range}")
        
        wb = load_workbook(file_path, data_only=True)
        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet}' not found"}
        
        ws = wb[sheet]
        data = []
        
        for row in ws[range]:
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        wb.close()
        
        logger.info(f"   ✅ Read {len(data)} rows")
        return {"success": True, "data": data, "range": range, "sheet": sheet}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


# ============================================================================
# WRITE TOOLS - Modify cells and data
# ============================================================================

@register_tool(
    name="WRITE_CELL",
    description="Write a value to a single cell",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "cell": "Cell address (e.g., 'B5')",
        "value": "Value to write (number, string, formula)"
    }
)
def write_cell(file_path: str, sheet: str, cell: str, value: Any) -> Dict[str, Any]:
    """
    Write a value to a single cell using XLWINGS (live Excel interaction)
    
    ✅ SAFE: Creates an edited copy (if not already edited), leaves template untouched!
    - Template: /path/to/Template.xlsx (UNCHANGED)
    - Edited:   /path/to/Template_edited_YYYYMMDD_HHMMSS.xlsx (MODIFIED)
    
    ⚙️ Strategy: xlwings for writing (preserves macros, formulas, live calculation)
    """
    try:
        logger.info(f"✏️ WRITE_CELL: {file_path} | {sheet} | {cell} = {value}")
        
        # Check if this is already an edited file
        if "_edited_" in file_path:
            edited_path = file_path
            template_path = file_path
            logger.info(f"   📂 Using existing edited file")
        else:
            # Create edited copy (leaves template untouched)
            edited_path = _create_edited_copy(file_path)
            if not edited_path:
                return {"success": False, "error": "Failed to create edited copy"}
            template_path = file_path
        
        # Use xlwings for WRITING (live Excel, preserves everything)
        import xlwings as xw
        logger.info("   🔧 Using xlwings for write operation (live Excel)")
        
        # Convert value to proper type
        if isinstance(value, str):
            try:
                if '.' in value:
                    value = float(value)
                else:
                    value = int(value)
            except (ValueError, AttributeError):
                pass
        
        # Open with xlwings (invisible Excel instance)
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(edited_path)
        
        try:
            ws = wb.sheets[sheet]
            ws.range(cell).value = value
            wb.save(edited_path)
            logger.info(f"   ✅ Written with xlwings: {cell} = {value}")
            logger.info(f"   📁 File: {edited_path}")
            
            wb.close()
            app.quit()
            
            return {
                "success": True, 
                "cell": cell, 
                "value": value, 
                "sheet": sheet,
                "template_file": template_path,
                "edited_file": edited_path
            }
            
        except Exception as e:
            logger.error(f"   ❌ xlwings error: {e}")
            wb.close()
            app.quit()
            return {"success": False, "error": f"xlwings failed: {str(e)}"}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "error": str(e)}


@register_tool(
    name="WRITE_CELLS",
    description="Write multiple cell values at once",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "cell_values": "Dict of cell:value pairs (e.g., {'B5': 10, 'C7': 20})"
    }
)
def write_cells(file_path: str, sheet: str, cell_values: Dict[str, Any]) -> Dict[str, Any]:
    """
    Write multiple cells at once using XLWINGS (live Excel interaction)
    
    ⚙️ Strategy: xlwings for writing (preserves macros, formulas, triggers recalc)
    """
    try:
        logger.info(f"✏️ WRITE_CELLS: {file_path} | {sheet} | {len(cell_values)} cells")
        
        # Check if this is already an edited file
        if "_edited_" in file_path:
            edited_path = file_path
            template_path = file_path
            logger.info(f"   📂 Using existing edited file")
        else:
            # Create edited copy (leaves template untouched)
            edited_path = _create_edited_copy(file_path)
            if not edited_path:
                return {"success": False, "error": "Failed to create edited copy"}
            template_path = file_path
        
        # Use xlwings for WRITING (live Excel)
        import xlwings as xw
        logger.info("   🔧 Using xlwings for batch write operation")
        
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(edited_path)
        
        try:
            ws = wb.sheets[sheet]
            
            for cell, value in cell_values.items():
                # Convert value to proper type
                if isinstance(value, str):
                    try:
                        if '.' in value:
                            value = float(value)
                        else:
                            value = int(value)
                    except (ValueError, AttributeError):
                        pass
                
                ws.range(cell).value = value
                logger.info(f"   ✏️ {cell} = {value}")
            
            wb.save(edited_path)
            logger.info(f"   ✅ Written {len(cell_values)} cells")
            logger.info(f"   📁 File: {edited_path}")
            
            wb.close()
            app.quit()
            
            return {
                "success": True, 
                "cells_written": len(cell_values), 
                "sheet": sheet,
                "template_file": template_path,
                "edited_file": edited_path
            }
            
        except Exception as e:
            logger.error(f"   ❌ xlwings error: {e}")
            wb.close()
            app.quit()
            return {"success": False, "error": f"xlwings failed: {str(e)}"}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {"success": False, "error": str(e)}


@register_tool(
    name="APPEND_ROW",
    description="Append a new row to a table",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "header_row": "Row number where headers are located",
        "row_values": "Dict of column_name:value pairs"
    }
)
def append_row(file_path: str, sheet: str, header_row: int, 
               row_values: Dict[str, Any]) -> Dict[str, Any]:
    """Append a new row to a table"""
    try:
        logger.info(f"➕ APPEND_ROW: {file_path} | {sheet}")
        
        # Read existing table
        df = pd.read_excel(file_path, sheet_name=sheet, header=header_row)
        
        # Append new row
        new_row = pd.DataFrame([row_values])
        df = pd.concat([df, new_row], ignore_index=True)
        
        # Write back
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', 
                           if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet, index=False, startrow=header_row)
        
        logger.info(f"   ✅ Row appended. New row count: {len(df)}")
        return {"success": True, "new_row_count": len(df), "sheet": sheet}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


# ============================================================================
# CALCULATE TOOLS - Trigger computations
# ============================================================================

@register_tool(
    name="RECALC",
    description="Recalculate formulas in the workbook (requires xlwings)",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Optional: Specific sheet to recalculate (None = all)"
    }
)
def recalc(file_path: str, sheet: Optional[str] = None) -> Dict[str, Any]:
    """Recalculate formulas using xlwings"""
    try:
        if not XLWINGS_AVAILABLE:
            return {"success": False, "error": "xlwings not available"}
        
        logger.info(f"🔄 RECALC: {file_path} | {sheet or 'all sheets'}")
        
        app = xw.App(visible=False)
        wb = app.books.open(file_path)
        
        if sheet:
            if sheet in [s.name for s in wb.sheets]:
                wb.sheets[sheet].api.Calculate()
            else:
                wb.close()
                app.quit()
                return {"success": False, "error": f"Sheet '{sheet}' not found"}
        else:
            wb.api.Calculate()
        
        wb.save()
        wb.close()
        app.quit()
        
        logger.info(f"   ✅ Recalculation complete")
        return {"success": True, "recalculated": sheet or "all"}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


# ============================================================================
# STRUCTURE TOOLS - Manage sheets and workbook structure
# ============================================================================

@register_tool(
    name="CREATE_SHEET",
    description="Create a new sheet in the workbook",
    parameters={
        "file_path": "Path to Excel file",
        "sheet_name": "Name for the new sheet"
    }
)
def create_sheet(file_path: str, sheet_name: str) -> Dict[str, Any]:
    """Create a new sheet"""
    try:
        logger.info(f"➕ CREATE_SHEET: {file_path} | {sheet_name}")
        
        wb = load_workbook(file_path)
        
        if sheet_name in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Sheet '{sheet_name}' already exists"}
        
        wb.create_sheet(sheet_name)
        wb.save(file_path)
        wb.close()
        
        logger.info(f"   ✅ Sheet created: {sheet_name}")
        return {"success": True, "sheet_name": sheet_name}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


@register_tool(
    name="COPY_SHEET",
    description="Copy a sheet to a new name (e.g., copy template)",
    parameters={
        "file_path": "Path to Excel file",
        "source_sheet": "Sheet to copy from",
        "target_sheet": "Name for the new sheet"
    }
)
def copy_sheet(file_path: str, source_sheet: str, target_sheet: str) -> Dict[str, Any]:
    """Copy a sheet"""
    try:
        logger.info(f"📋 COPY_SHEET: {file_path} | {source_sheet} → {target_sheet}")
        
        wb = load_workbook(file_path)
        
        if source_sheet not in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Source sheet '{source_sheet}' not found"}
        
        if target_sheet in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"Target sheet '{target_sheet}' already exists"}
        
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = target_sheet
        
        wb.save(file_path)
        wb.close()
        
        logger.info(f"   ✅ Sheet copied: {source_sheet} → {target_sheet}")
        return {"success": True, "source": source_sheet, "target": target_sheet}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


@register_tool(
    name="LIST_SHEETS",
    description="List all sheets in the workbook",
    parameters={
        "file_path": "Path to Excel file"
    }
)
def list_sheets(file_path: str) -> Dict[str, Any]:
    """List all sheets"""
    try:
        logger.info(f"📋 LIST_SHEETS: {file_path}")
        
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        
        logger.info(f"   ✅ Found {len(sheets)} sheets: {', '.join(sheets[:5])}...")
        return {"success": True, "sheets": sheets, "count": len(sheets)}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


# ============================================================================
# ANALYSIS TOOLS - Understand content
# ============================================================================

@register_tool(
    name="FIND_PARAMETER",
    description="Find a parameter by label (e.g., find 'Span' label and return its cell)",
    parameters={
        "file_path": "Path to Excel file",
        "sheet": "Sheet name",
        "label": "Label to search for (e.g., 'Span', 'Load')",
        "search_area": "Optional: Cell range to search in (e.g., 'A1:B50')"
    }
)
def find_parameter(file_path: str, sheet: str, label: str, 
                   search_area: Optional[str] = None) -> Dict[str, Any]:
    """
    Find a parameter by its label with FUZZY matching
    
    Handles variations like:
    - 'span' matches 'Span', 'SPAN', 'span length', 'Overhang n', etc.
    - 'load' matches 'Load', 'Dead Load', 'Live Load', etc.
    """
    try:
        logger.info(f"🔍 FIND_PARAMETER: {file_path} | {sheet} | '{label}'")
        
        wb = load_workbook(file_path, data_only=True)
        if sheet not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet}' not found"}
        
        ws = wb[sheet]
        
        # Normalize search term - extract key words
        label_lower = label.lower()
        # Remove common suffixes and underscores
        label_clean = label_lower.replace('_', ' ').replace('-', ' ')
        key_words = [w for w in label_clean.split() if len(w) > 2]  # Words > 2 chars
        
        logger.info(f"   🔍 Searching for keywords: {key_words}")
        
        found_cells = []
        
        max_row = ws.max_row if not search_area else 100
        max_col = ws.max_column if not search_area else 10
        
        for row in range(1, min(max_row + 1, 200)):
            for col in range(1, min(max_col + 1, 20)):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    cell_lower = cell.value.lower()
                    
                    # Fuzzy match: check if ANY keyword is in the cell
                    matches = any(keyword in cell_lower for keyword in key_words)
                    
                    if matches:
                        # Found label, check if next cell has a value
                        value_cell = ws.cell(row, col + 1)
                        found_cells.append({
                            "label_cell": f"{get_column_letter(col)}{row}",
                            "value_cell": f"{get_column_letter(col + 1)}{row}",
                            "label_text": cell.value,
                            "current_value": value_cell.value
                        })
        
        wb.close()
        
        if found_cells:
            logger.info(f"   ✅ Found {len(found_cells)} matches")
            return {"success": True, "matches": found_cells, "count": len(found_cells)}
        else:
            logger.info(f"   ⚠️ No matches found for '{label}'")
            return {"success": True, "matches": [], "count": 0}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


# ============================================================================
# SAFETY TOOLS - Backup and validation
# ============================================================================

@register_tool(
    name="CREATE_BACKUP",
    description="Create a backup copy of the file before making changes",
    parameters={
        "file_path": "Path to Excel file"
    }
)
def create_backup(file_path: str) -> Dict[str, Any]:
    """Create a backup copy"""
    try:
        logger.info(f"💾 CREATE_BACKUP: {file_path}")
        
        file_path = Path(file_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.parent / f"{file_path.stem}_backup_{timestamp}{file_path.suffix}"
        
        shutil.copy2(file_path, backup_path)
        
        logger.info(f"   ✅ Backup created: {backup_path.name}")
        return {"success": True, "backup_path": str(backup_path)}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


@register_tool(
    name="VALIDATE_FILE",
    description="Validate that a file exists and can be opened",
    parameters={
        "file_path": "Path to Excel file"
    }
)
def validate_file(file_path: str) -> Dict[str, Any]:
    """Validate a file"""
    try:
        logger.info(f"✅ VALIDATE_FILE: {file_path}")
        
        file_path = Path(file_path)
        
        if not file_path.exists():
            return {"success": False, "error": "File does not exist"}
        
        if not file_path.suffix in ['.xlsx', '.xlsm']:
            return {"success": False, "error": "Not an Excel file"}
        
        # Try to open it
        wb = load_workbook(file_path, read_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
        
        logger.info(f"   ✅ Valid Excel file with {sheet_count} sheets")
        return {"success": True, "sheet_count": sheet_count, "file_size": file_path.stat().st_size}
        
    except Exception as e:
        logger.error(f"   ❌ Error: {e}")
        return {"success": False, "error": str(e)}


# ============================================================================
# TOOL EXECUTION ENGINE
# ============================================================================

def execute_tool(tool_name: str, **kwargs) -> Dict[str, Any]:
    """
    Execute a tool by name with given parameters
    
    This is the main interface the LLM uses to call tools
    """
    if tool_name not in TOOL_REGISTRY:
        return {"success": False, "error": f"Tool '{tool_name}' not found"}
    
    tool_info = TOOL_REGISTRY[tool_name]
    tool_function = tool_info["function"]
    
    try:
        result = tool_function(**kwargs)
        return result
    except Exception as e:
        logger.error(f"Tool execution failed: {tool_name} - {e}")
        return {"success": False, "error": str(e)}


def get_tool_list() -> List[Dict[str, Any]]:
    """
    Get list of all available tools with their descriptions
    
    The LLM uses this to know what tools are available
    """
    tools = []
    for name, info in TOOL_REGISTRY.items():
        tools.append({
            "name": name,
            "description": info["description"],
            "parameters": info["parameters"],
            "category": info.get("category", "GENERAL")
        })
    return tools


def get_tool_documentation() -> str:
    """
    Get formatted documentation for all tools
    
    This can be included in the LLM prompt
    """
    doc_lines = ["# Excel Agent Tools", "", "Available tools for spreadsheet operations:", ""]
    
    for name, info in TOOL_REGISTRY.items():
        doc_lines.append(f"## {name}")
        doc_lines.append(f"**Description**: {info['description']}")
        doc_lines.append("**Parameters**:")
        for param_name, param_desc in info["parameters"].items():
            doc_lines.append(f"  - `{param_name}`: {param_desc}")
        doc_lines.append("")
    
    return "\n".join(doc_lines)


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == "__main__":
    print("🔧 Excel Agent Tool System")
    print("=" * 60)
    print(f"Available tools: {len(TOOL_REGISTRY)}")
    print()
    
    # Print tool documentation
    print(get_tool_documentation())
    
    # Example: List tools for LLM
    print("\n" + "=" * 60)
    print("Tools available to LLM:")
    print(json.dumps(get_tool_list(), indent=2))

